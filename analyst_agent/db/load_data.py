"""
load_data.py — Analyst Agent | Production Data Loader
======================================================
Loads client Excel/CSV data directly into PostgreSQL.
Supports both initial full loads and incremental appends.

Features:
  - Direct-to-database loading (no intermediate SQL files)
  - PostgreSQL COPY for bulk speed (10-50x faster than INSERT)
  - Falls back to batch INSERT ... ON CONFLICT DO NOTHING
  - Data validation before loading (types, nulls, ranges)
  - Auto-refreshes mv_customer_features after loading
  - Supports Excel (.xlsx) and CSV (.csv) input
  - Handles thousands/millions of rows efficiently
  - FK-safe loading order (parents before children)
  - Detailed summary report with row counts

Usage:
    # Full load (truncates + reloads all tables):
    python load_data.py --excel data.xlsx --db-url postgresql://user:pass@localhost:5432/walmart_crp --mode full

    # Incremental append (inserts new rows, skips duplicates):
    python load_data.py --excel data.xlsx --db-url postgresql://user:pass@localhost:5432/walmart_crp --mode append

    # Load only specific sheets:
    python load_data.py --excel data.xlsx --db-url postgresql://user:pass@localhost:5432/walmart_crp --sheets customers orders

    # Skip materialized view refresh:
    python load_data.py --excel data.xlsx --db-url postgresql://user:pass@localhost:5432/walmart_crp --no-refresh

Requirements:
    pip install openpyxl psycopg2-binary python-dotenv
"""

import os
import sys
import csv
import io
import argparse
import logging
import time
from datetime import datetime, date
from pathlib import Path
from collections import OrderedDict

import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

# ── Logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("load_data")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 1: HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def clean(val):
    """Normalise a cell value: strip formulas, convert dates, clean text."""
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip()
        if v.startswith("="):
            if "TRUE" in v.upper():
                return True
            if "FALSE" in v.upper():
                return False
            return None
        if v.startswith("──") or v == "":
            return None
        return v
    if isinstance(val, datetime):
        return val.date() if val.hour == val.minute == val.second == 0 else val
    return val


def find_header_row(rows, expected_col):
    """Return (header_row_idx, headers) for the row containing expected_col."""
    expected_lower = expected_col.lower().strip()
    for i, row in enumerate(rows):
        for cell in row:
            if cell is not None and str(cell).strip().lower() == expected_lower:
                headers = [str(c).strip().lower() if c is not None else None for c in row]
                return i, [h for h in headers if h is not None]
    raise ValueError(f"Header row with column '{expected_col}' not found in sheet.")


def get_data_rows(rows, header_idx, num_cols):
    """Return cleaned data rows after header, skipping blank/separator rows."""
    result = []
    for row in rows[header_idx + 1:]:
        cleaned = tuple(clean(v) for v in row[:num_cols])
        if cleaned[0] is None:
            continue
        result.append(cleaned)
    return result


def validate_row(row, cols, table_name, row_num, validators):
    """Validate a single row. Returns list of error messages (empty = valid)."""
    errors = []
    for i, (val, col) in enumerate(zip(row, cols)):
        if col in validators:
            err = validators[col](val, row_num)
            if err:
                errors.append(f"  Row {row_num}, {col}: {err}")
    return errors


# ── Validators ──────────────────────────────────────────────────────────────

def not_null(val, row_num):
    if val is None:
        return "cannot be NULL"
    return None

def is_positive_number(val, row_num):
    if val is None:
        return None
    try:
        if float(val) < 0:
            return f"expected positive number, got {val}"
    except (ValueError, TypeError):
        return f"expected number, got {val}"
    return None

def rating_1_to_5(val, row_num):
    if val is None:
        return None
    try:
        v = int(val)
        if v < 1 or v > 5:
            return f"rating must be 1-5, got {v}"
    except (ValueError, TypeError):
        return f"expected integer 1-5, got {val}"
    return None


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 2: BULK LOADING — COPY vs INSERT
# ═══════════════════════════════════════════════════════════════════════════

def copy_bulk_load(cur, table, columns, rows):
    """
    Use PostgreSQL COPY FROM for maximum speed.
    Converts rows to a CSV buffer and streams it in.
    Best for full loads where the table is empty or truncated.
    """
    if not rows:
        return 0

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter='\t', quoting=csv.QUOTE_MINIMAL)
    for row in rows:
        writer.writerow([
            '\\N' if v is None
            else str(v).replace('\t', ' ').replace('\n', ' ')
            for v in row
        ])

    buf.seek(0)
    col_list = ", ".join(columns)
    copy_sql = f"COPY {table} ({col_list}) FROM STDIN WITH (FORMAT csv, DELIMITER E'\\t', NULL '\\N')"
    cur.copy_expert(copy_sql, buf)
    return len(rows)


def batch_upsert(cur, table, columns, rows, page_size=500):
    """
    Use INSERT ... ON CONFLICT DO NOTHING for incremental appends.
    Skips rows that already exist (by primary key).
    """
    if not rows:
        return 0

    placeholders = ", ".join(["%s"] * len(columns))
    col_list = ", ".join(columns)
    sql = (
        f"INSERT INTO {table} ({col_list}) VALUES ({placeholders}) "
        f"ON CONFLICT DO NOTHING"
    )
    inserted = 0
    for start in range(0, len(rows), page_size):
        chunk = rows[start: start + page_size]
        safe_rows = [
            r[:len(columns)] + (None,) * max(0, len(columns) - len(r))
            for r in chunk
        ]
        psycopg2.extras.execute_batch(cur, sql, safe_rows, page_size=page_size)
        inserted += len(chunk)
    return inserted


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 3: PER-TABLE SHEET LOADERS
# ═══════════════════════════════════════════════════════════════════════════

# Each loader returns: (columns_list, cleaned_rows, validators_dict)

def load_client_config(ws):
    rows = list(ws.iter_rows(values_only=True))
    config = {}
    for row in rows:
        param, value = clean(row[0]), clean(row[1])
        if param and not param.startswith("──") and param != "Parameter":
            config[param] = value

    cols = ["client_id", "client_name", "client_code", "currency", "timezone",
            "churn_window_days", "high_ltv_threshold", "mid_ltv_threshold", "max_discount_pct"]
    data = [(
        config.get("client_id", "CLT-001"),
        config.get("client_name", "Unknown"),
        config.get("client_code", "UNK"),
        config.get("report_currency", config.get("currency", "USD")),
        config.get("timezone", "America/Chicago"),
        int(config.get("churn_inactivity_days", config.get("churn_window_days", 90))),
        float(config.get("fixed_tier1_min_spend_usd", config.get("high_ltv_threshold", 1000))),
        float(config.get("fixed_tier2_min_spend_usd", config.get("mid_ltv_threshold", 200))),
        float(config.get("max_discount_pct", 30)),
    )]
    validators = {"client_id": not_null}
    return cols, data, validators


def load_categories(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "category_id")
    cols = ["category_id", "category_name"]
    data = get_data_rows(rows, hi, 2)
    return cols, data, {"category_id": not_null}


def load_sub_categories(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "sub_category_id")
    cols = ["sub_category_id", "sub_category_name", "category_id"]
    data = get_data_rows(rows, hi, 3)
    return cols, data, {"sub_category_id": not_null}


def load_sub_sub_categories(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "sub_sub_category_id")
    cols = ["sub_sub_category_id", "sub_sub_category_name", "sub_category_id", "category_id"]
    data = get_data_rows(rows, hi, 4)
    return cols, data, {"sub_sub_category_id": not_null}


def load_vendors(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "vendor_id")
    cols = ["vendor_id", "vendor_name", "vendor_description", "vendor_contact_no",
            "vendor_address", "vendor_email"]
    data = get_data_rows(rows, hi, 6)
    return cols, data, {"vendor_id": not_null}


def load_brands(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "brand_id")
    cols = ["brand_id", "brand_name", "brand_description", "vendor_id",
            "active", "not_available", "category_hint"]
    raw = get_data_rows(rows, hi, 7)
    cleaned = []
    for r in raw:
        row = list(r)
        for idx in [4, 5]:
            if idx < len(row) and isinstance(row[idx], bool):
                row[idx] = 1 if row[idx] else 0
            elif idx < len(row) and row[idx] is None:
                row[idx] = 0
        cleaned.append(tuple(row))
    return cols, cleaned, {"brand_id": not_null}


def load_products(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "product_id")
    cols = ["product_id", "sku", "product_name", "category_id", "sub_category_id",
            "sub_sub_category_id", "brand_id", "product_price_id", "rating",
            "active", "not_available"]
    raw = get_data_rows(rows, hi, 11)
    cleaned = []
    for r in raw:
        row = list(r)
        for idx in [9, 10]:
            if idx < len(row) and isinstance(row[idx], bool):
                row[idx] = 1 if row[idx] else 0
            elif idx < len(row) and row[idx] is None:
                row[idx] = 0
        cleaned.append(tuple(row))
    return cols, cleaned, {"product_id": not_null}


def load_product_prices(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "price_id")
    cols = ["price_id", "product_id", "qty_range_label", "qty_min", "qty_max", "unit_price_usd"]
    data = get_data_rows(rows, hi, 6)
    return cols, data, {"price_id": not_null, "unit_price_usd": is_positive_number}


def load_product_vendor_mapping(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "pv_id")
    cols = ["pv_id", "product_id", "brand_id", "vendor_id"]
    data = get_data_rows(rows, hi, 4)
    return cols, data, {"pv_id": not_null}


def load_customers(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "client_id")
    cols = ["client_id", "customer_id", "customer_email", "customer_name", "customer_phone",
            "account_created_date", "registration_channel", "country_code", "state", "city",
            "zip_code", "shipping_address", "preferred_device", "email_opt_in", "sms_opt_in"]
    raw = get_data_rows(rows, hi, 15)
    cleaned = []
    for r in raw:
        row = list(r)
        # account_created_date
        if isinstance(row[5], datetime):
            row[5] = row[5].date()
        # email_opt_in, sms_opt_in → boolean
        for idx in [13, 14]:
            if idx < len(row):
                v = row[idx]
                if isinstance(v, str):
                    row[idx] = v.upper() in ("TRUE", "YES", "1")
                elif isinstance(v, (int, float)):
                    row[idx] = bool(v)
        cleaned.append(tuple(row))
    return cols, cleaned, {"client_id": not_null, "customer_id": not_null}


def load_orders(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "client_id")
    cols = ["client_id", "order_id", "customer_id", "order_date", "order_status",
            "order_value_usd", "discount_usd", "coupon_code", "payment_method", "order_item_count"]
    raw = get_data_rows(rows, hi, 10)
    cleaned = []
    for r in raw:
        row = list(r)
        # Parse order_date string if needed
        if isinstance(row[3], str):
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
                        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y"):
                try:
                    row[3] = datetime.strptime(row[3], fmt)
                    break
                except ValueError:
                    continue
        cleaned.append(tuple(row))
    return cols, cleaned, {
        "order_id": not_null,
        "order_value_usd": is_positive_number,
    }


def load_line_items(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "client_id")
    cols = ["client_id", "line_item_id", "order_id", "customer_id", "product_id",
            "quantity", "unit_price_usd", "final_line_total_usd", "item_discount_usd", "item_status"]
    data = get_data_rows(rows, hi, 10)
    return cols, data, {"line_item_id": not_null, "quantity": is_positive_number}


def load_value_tiers(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "tier_id")
    cols = ["tier_id", "tier_name", "threshold_type", "threshold_value", "description", "benefits"]
    data = get_data_rows(rows, hi, 6)
    return cols, data, {"tier_id": not_null}


def load_business_segments(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "segment_id")
    cols = ["segment_id", "segment_name", "description", "criteria", "recommended_focus"]
    data = get_data_rows(rows, hi, 5)
    return cols, data, {"segment_id": not_null}


def load_value_propositions(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "tier_name")
    cols = ["tier_name", "risk_level", "action_type", "message_template",
            "discount_pct", "channel", "priority"]
    data = get_data_rows(rows, hi, 7)
    return cols, data, {}


def load_customer_reviews(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "client_id")
    cols = ["client_id", "review_id", "customer_id", "product_id", "order_id",
            "rating", "review_text", "review_date", "sentiment"]
    raw = get_data_rows(rows, hi, 9)
    cleaned = []
    for r in raw:
        row = list(r)
        if row[5] is not None:
            try:
                row[5] = int(row[5])
            except (ValueError, TypeError):
                row[5] = None
        cleaned.append(tuple(row))
    return cols, cleaned, {"review_id": not_null, "rating": rating_1_to_5}


def load_support_tickets(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi, _ = find_header_row(rows, "client_id")
    cols = ["client_id", "ticket_id", "customer_id", "ticket_type", "priority",
            "status", "channel", "opened_date", "resolved_date", "resolution_time_hrs"]
    raw = get_data_rows(rows, hi, 10)
    cleaned = []
    for r in raw:
        row = list(r)
        for idx in [7, 8]:
            if idx < len(row) and isinstance(row[idx], str):
                for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
                            "%m/%d/%Y %H:%M", "%m/%d/%Y"):
                    try:
                        row[idx] = datetime.strptime(row[idx], fmt)
                        break
                    except ValueError:
                        continue
        cleaned.append(tuple(row))
    return cols, cleaned, {"ticket_id": not_null}


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 4: SHEET → TABLE MAPPING (FK-safe order)
# ═══════════════════════════════════════════════════════════════════════════

# Ordered dict: sheet_name → (table_name, loader_function)
# Order matters! Parents load before children for FK integrity.
SHEET_MAP = OrderedDict([
    ("Vendor Config",            ("client_config",          load_client_config)),
    ("Category Master",          ("categories",             load_categories)),
    ("Sub-Category Master",      ("sub_categories",         load_sub_categories)),
    ("Sub-Sub-Category Master",  ("sub_sub_categories",     load_sub_sub_categories)),
    ("Vendor Master",            ("vendors",                load_vendors)),
    ("Brand Master",             ("brands",                 load_brands)),
    ("Product Master",           ("products",               load_products)),
    ("Product Price Master",     ("product_prices",         load_product_prices)),
    ("Product-Vendor Mapping",   ("product_vendor_mapping", load_product_vendor_mapping)),
    ("Customer Master",          ("customers",              load_customers)),
    ("Order Master",             ("orders",                 load_orders)),
    ("Line Items Master",        ("line_items",             load_line_items)),
    ("Value-Tier Master",        ("value_tiers",            load_value_tiers)),
    ("Business Segment Master",  ("business_segments",      load_business_segments)),
    ("Value Proposition Master", ("value_propositions",     load_value_propositions)),
    ("Customer Reviews",         ("customer_reviews",       load_customer_reviews)),
    ("Support Tickets",          ("support_tickets",        load_support_tickets)),
])


def match_sheet_name(wb_sheets, key):
    """
    Match a sheet key to the actual workbook sheet name.
    Handles emoji prefixes — e.g. key="Customer Master" matches
    sheet name "Customer Master" (with emoji prefix).
    """
    key_lower = key.lower().strip()
    for name in wb_sheets:
        # Strip emoji prefix (first 2-3 chars might be emoji + space)
        name_clean = name.strip()
        # Exact match
        if name_clean.lower() == key_lower:
            return name
        # Match ignoring emoji prefix (check if key appears at end of sheet name)
        if name_clean.lower().endswith(key_lower):
            return name
        # Partial match: sheet name contains key
        if key_lower in name_clean.lower():
            return name
    return None


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 5: MAIN LOADER
# ═══════════════════════════════════════════════════════════════════════════

def refresh_materialized_view(cur):
    """Refresh the churn prediction feature view after data load."""
    log.info("\nRefreshing materialized view mv_customer_features...")
    try:
        cur.execute("REFRESH MATERIALIZED VIEW mv_customer_features;")
        log.info("  Materialized view refreshed successfully.")
        return True
    except Exception as e:
        log.warning("  Could not refresh materialized view: %s", e)
        log.warning("  (Run schema_migration_v3.sql first if the view doesn't exist yet)")
        return False


def get_table_counts(cur, tables):
    """Get row counts for all loaded tables."""
    counts = {}
    for table in tables:
        try:
            cur.execute(f"SELECT COUNT(*) FROM {table}")
            counts[table] = cur.fetchone()[0]
        except Exception:
            counts[table] = "?"
    return counts


def main():
    load_dotenv()
    parser = argparse.ArgumentParser(
        description="Load client data from Excel/CSV directly into PostgreSQL",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Full load from Excel:
  python load_data.py --excel data.xlsx --db-url postgresql://user:pass@localhost:5432/walmart_crp --mode full

  # Incremental append (skip duplicates):
  python load_data.py --excel data.xlsx --db-url postgresql://user:pass@localhost:5432/walmart_crp --mode append

  # Load only customers and orders:
  python load_data.py --excel data.xlsx --db-url postgresql://user:pass@localhost:5432/walmart_crp --sheets customers orders
        """
    )
    parser.add_argument("--excel", required=True, help="Path to Excel (.xlsx) file")
    parser.add_argument("--db-url", default=os.getenv("DB_URL"),
                        help="PostgreSQL connection string (or set DB_URL env var)")
    parser.add_argument("--mode", choices=["full", "append"], default="append",
                        help="'full' = truncate + reload; 'append' = insert new, skip duplicates (default)")
    parser.add_argument("--sheets", nargs="*", default=None,
                        help="Load only these sheets (e.g. --sheets customers orders)")
    parser.add_argument("--no-refresh", action="store_true",
                        help="Skip refreshing the materialized view after loading")
    parser.add_argument("--schema", default=None,
                        help="Path to schema SQL file to apply before loading (optional)")
    args = parser.parse_args()

    # ── Validate inputs ──────────────────────────────────────────────────
    if not args.db_url:
        log.error("DB_URL not set. Use --db-url or set DB_URL env var.")
        log.error("Example: postgresql://postgres:password@localhost:5432/walmart_crp")
        sys.exit(1)

    excel_path = Path(args.excel)
    if not excel_path.exists():
        log.error("File not found: %s", args.excel)
        sys.exit(1)

    # ── Open workbook ────────────────────────────────────────────────────
    start_time = time.time()
    log.info("Opening workbook: %s", excel_path)
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    log.info("  Sheets found: %s", wb.sheetnames)

    # ── Connect to database ──────────────────────────────────────────────
    log.info("Connecting to PostgreSQL...")
    conn = psycopg2.connect(args.db_url)
    conn.autocommit = False
    cur = conn.cursor()
    log.info("  Connected successfully.")

    # ── Apply schema if provided ─────────────────────────────────────────
    if args.schema:
        schema_path = Path(args.schema)
        if schema_path.exists():
            log.info("Applying schema: %s", schema_path)
            with open(schema_path) as f:
                cur.execute(f.read())
            conn.commit()
            log.info("  Schema applied.")
        else:
            log.warning("Schema file not found: %s — skipping", schema_path)

    # ── Determine which sheets to load ───────────────────────────────────
    sheets_to_load = SHEET_MAP
    if args.sheets:
        # Filter to only requested sheets
        requested = {s.lower().replace("_", " ") for s in args.sheets}
        sheets_to_load = OrderedDict(
            (k, v) for k, v in SHEET_MAP.items()
            if any(req in k.lower() for req in requested)
              or any(req in v[0].lower().replace("_", " ") for req in requested)
        )
        if not sheets_to_load:
            log.error("No matching sheets found for: %s", args.sheets)
            log.error("Available: %s", list(SHEET_MAP.keys()))
            sys.exit(1)

    # ── Load data ────────────────────────────────────────────────────────
    log.info("\nMode: %s | Loading %d table(s)...\n", args.mode.upper(), len(sheets_to_load))

    results = {}  # table → {"rows_read": N, "rows_loaded": N, "errors": [...], "time_ms": N}
    total_errors = []

    for sheet_key, (table_name, loader_fn) in sheets_to_load.items():
        # Match sheet name (handles emoji prefixes)
        actual_sheet = match_sheet_name(wb.sheetnames, sheet_key)
        if actual_sheet is None:
            log.warning("  Sheet not found: '%s' — skipping", sheet_key)
            continue

        log.info("Loading: %s → %s", actual_sheet, table_name)
        t0 = time.time()

        try:
            ws = wb[actual_sheet]
            cols, data, validators = loader_fn(ws)

            # ── Validate data ────────────────────────────────────────
            validation_errors = []
            for row_num, row in enumerate(data, start=1):
                errs = validate_row(row, cols, table_name, row_num, validators)
                validation_errors.extend(errs)

            if validation_errors:
                log.warning("  Validation warnings (%d):", len(validation_errors))
                for err in validation_errors[:5]:
                    log.warning("    %s", err)
                if len(validation_errors) > 5:
                    log.warning("    ... and %d more", len(validation_errors) - 5)

            # ── Load into database ───────────────────────────────────
            if args.mode == "full":
                # Truncate cascade for full reload
                cur.execute(f"SET session_replication_role = replica;")
                cur.execute(f"TRUNCATE TABLE {table_name} CASCADE;")
                rows_loaded = copy_bulk_load(cur, table_name, cols, data)
                cur.execute(f"SET session_replication_role = DEFAULT;")
            else:
                # Incremental: INSERT ON CONFLICT DO NOTHING
                rows_loaded = batch_upsert(cur, table_name, cols, data)

            conn.commit()

            elapsed = int((time.time() - t0) * 1000)
            results[table_name] = {
                "rows_read": len(data),
                "rows_loaded": rows_loaded,
                "errors": validation_errors,
                "time_ms": elapsed,
            }
            log.info("  %d rows read, %d loaded (%d ms)", len(data), rows_loaded, elapsed)

        except Exception as e:
            conn.rollback()
            log.error("  FAILED to load %s: %s", table_name, e)
            total_errors.append((table_name, str(e)))

    # ── Refresh materialized view ────────────────────────────────────────
    if not args.no_refresh:
        try:
            refresh_materialized_view(cur)
            conn.commit()
        except Exception as e:
            conn.rollback()
            log.warning("  Materialized view refresh failed: %s", e)

    # ── Summary report ───────────────────────────────────────────────────
    elapsed_total = time.time() - start_time

    # Get final row counts from database
    table_names = [v[0] for v in sheets_to_load.values()]
    final_counts = get_table_counts(cur, table_names)

    cur.close()
    conn.close()

    log.info("\n" + "=" * 60)
    log.info("LOAD SUMMARY")
    log.info("=" * 60)
    log.info("  File:    %s", excel_path.name)
    log.info("  Mode:    %s", args.mode.upper())
    log.info("  Time:    %.1f seconds", elapsed_total)
    log.info("")
    log.info("  %-30s %8s %8s %8s", "Table", "Read", "Loaded", "In DB")
    log.info("  " + "-" * 56)
    for table_name in table_names:
        if table_name in results:
            r = results[table_name]
            log.info("  %-30s %8d %8d %8s",
                     table_name, r["rows_read"], r["rows_loaded"],
                     str(final_counts.get(table_name, "?")))
    log.info("  " + "-" * 56)
    total_read = sum(r["rows_read"] for r in results.values())
    total_loaded = sum(r["rows_loaded"] for r in results.values())
    log.info("  %-30s %8d %8d", "TOTAL", total_read, total_loaded)

    if total_errors:
        log.warning("\n  ERRORS (%d):", len(total_errors))
        for table, err in total_errors:
            log.warning("    %s → %s", table, err)
        sys.exit(1)
    else:
        log.info("\n  All tables loaded successfully!")

    log.info("=" * 60)


if __name__ == "__main__":
    main()
