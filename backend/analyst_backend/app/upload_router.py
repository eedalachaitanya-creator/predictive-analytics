"""
upload_router.py — CSV/Excel file upload endpoints (BATCH FLOW)
===============================================================
POST   /api/v1/uploads/{masterType}  — Upload a file into the pending batch
GET    /api/v1/uploads               — List staged files in the pending batch
GET    /api/v1/uploads/batch         — Pending batch summary for a client
DELETE /api/v1/uploads/{masterType}  — Remove a file from the pending batch
POST   /api/v1/uploads/commit        — Validate + move staging → real tables
POST   /api/v1/uploads/discard       — Throw away the pending batch

HOW IT WORKS (NEW BATCH FLOW):
1. User picks a CSV/Excel file on the Upload page
2. Frontend sends it as multipart/form-data to POST /uploads/{masterType}
3. Backend reads + validates the file with pandas
4. File saved to disk (audit trail) + rows inserted into staging_<table>
   with a batch_id that groups them with other files the user uploads
5. User uploads more files — in any order, optional types can be skipped
   (re-uploading same master_type REPLACES its staging rows — last wins)
6. When done, user clicks "Commit Batch":
   - POST /uploads/commit runs pre-flight FK checks against staging + real
   - If clean, opens a transaction with SET CONSTRAINTS ALL DEFERRED
   - Moves all staging rows into real tables (DEFERRED FKs check at COMMIT)
   - Refreshes materialized view, clears staging, marks batch committed
7. If the user changes their mind: POST /uploads/discard clears the batch
"""

import os
import io
import csv
import re
import uuid
import logging
from datetime import datetime

import pandas as pd
from sqlalchemy import text
from fastapi import APIRouter, File, Form, UploadFile, Query, HTTPException, Header, Request, Depends, Response
from typing import Optional

from app.database import engine
from app.auth_router import _find_user_by_token, get_current_user
from app.audit_logger import log_audit_event


# Audit fix 2026-04-29 (#1): centralized tenant access check for every
# endpoint in this router. Same pattern used in chat_router. super_admin
# (clientAccess == ['*']) bypasses the per-tenant check. Anyone else
# must have the requested client_id in their clientAccess list.
def _require_client_access(user: dict, client_id: str) -> None:
    if user.get("role") == "super_admin" or "*" in (user.get("clientAccess") or []):
        return
    if client_id not in (user.get("clientAccess") or []):
        raise HTTPException(
            status_code=403,
            detail=f"You do not have access to client {client_id}",
        )


# Audit fix 2026-04-29 (#2): sanitize filenames before they hit disk.
# os.path.join doesn't reject `../` segments, so a raw UploadFile.filename
# can escape UPLOAD_DIR and write arbitrary paths. We strip everything
# that isn't a basic ASCII letter/digit/dash/underscore/dot, and then
# take basename() as belt-and-braces. Audit-relevant filenames get
# normalised to a safe slug; the original value is logged separately.
_SAFE_FILENAME_RE = re.compile(r"[^A-Za-z0-9._-]")


def _sanitize_filename(raw: str) -> str:
    """Return a path-traversal-safe filename derived from a user-supplied one."""
    if not raw:
        return "upload.dat"
    # basename() drops any leading path segments before the final /,
    # then we kill any remaining unsafe character.
    cleaned = _SAFE_FILENAME_RE.sub("_", os.path.basename(str(raw)))
    # Reject empty / dot-only results that would still be problematic.
    if not cleaned or cleaned in (".", ".."):
        return "upload.dat"
    return cleaned[:200]  # cap length so we can't blow path-name limits


# Audit fix 2026-04-29 (#3): cap file size at 50 MB to prevent OOM via
# pathological multi-GB uploads. The Excel files our pipeline expects
# are well under 5 MB; 50 MB is a generous ceiling. Override via env
# var if a tenant ever needs it raised.
MAX_UPLOAD_BYTES = int(os.environ.get("CRP_MAX_UPLOAD_BYTES", str(50 * 1024 * 1024)))

router = APIRouter(prefix="/api/v1", tags=["uploads"])
log = logging.getLogger("upload")

# ── Valid master types ───────────────────────────────────────────────────────
VALID_MASTER_TYPES = {
    "customer", "order", "line_items",
    "product", "price", "vendor_map",
    "category", "sub_category", "sub_sub_category",
    "brand", "vendor",
    "customer_reviews", "support_tickets",
    "login_event",
}

# Directory to save uploaded files
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "uploads")
try:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
except OSError:
    UPLOAD_DIR = "/tmp/crp_uploads"
    os.makedirs(UPLOAD_DIR, exist_ok=True)


# ── Master type → database table + expected columns ─────────────────────
# Maps each upload type to: (table_name, list_of_db_columns)
# The column list defines the ORDER columns should appear in the DB.
# If the CSV has different column names, we rename them to match.
MASTER_TYPE_TO_TABLE = {
    "customer": (
        "customers",
        # 16 columns — `last_login_date` was added 2026-04-24 for the
        # login-aware churn rule (migration 2026_04_24_last_login_date.sql).
        # Optional in the uploaded file; missing values land NULL.
        ["client_id", "customer_id", "customer_email", "customer_name",
         "customer_phone", "account_created_date", "registration_channel",
         "country_code", "state", "city", "zip_code", "shipping_address",
         "preferred_device", "email_opt_in", "sms_opt_in",
         "last_login_date"],
    ),
    "order": (
        "orders",
        ["client_id", "order_id", "customer_id", "order_date", "order_status",
         "order_value_usd", "discount_usd", "coupon_code", "payment_method",
         "order_item_count"],
    ),
    "line_items": (
        "line_items",
        ["client_id", "line_item_id", "order_id", "customer_id", "product_id",
         "quantity", "unit_price_usd", "final_line_total_usd",
         "item_discount_usd", "item_status"],
    ),
    "product": (
        "products",
        ["client_id", "product_id", "sku", "product_name", "category_id",
         "sub_category_id", "sub_sub_category_id", "brand_id",
         "product_price_id", "rating", "active", "not_available"],
    ),
    "price": (
        "product_prices",
        ["client_id", "price_id", "product_id", "qty_range_label", "qty_min",
         "qty_max", "unit_price_usd", "cost_price_usd"],
    ),
    "vendor_map": (
        "product_vendor_mapping",
        ["client_id", "pv_id", "product_id", "brand_id", "vendor_id"],
    ),
    "category": (
        "categories",
        ["client_id", "category_id", "category_name"],
    ),
    "sub_category": (
        "sub_categories",
        ["client_id", "sub_category_id", "sub_category_name", "category_id"],
    ),
    "sub_sub_category": (
        "sub_sub_categories",
        ["client_id", "sub_sub_category_id", "sub_sub_category_name",
         "sub_category_id", "category_id"],
    ),
    "brand": (
        "brands",
        ["client_id", "brand_id", "brand_name", "brand_description",
         "vendor_id", "active", "not_available", "category_hint"],
    ),
    "vendor": (
        "vendors",
        ["client_id", "vendor_id", "vendor_name", "vendor_description",
         "vendor_contact_no", "vendor_address", "vendor_email"],
    ),
    "customer_reviews": (
        "customer_reviews",
        ["client_id", "review_id", "customer_id", "product_id", "order_id",
         "rating", "review_text", "review_date", "sentiment", "source"],
    ),
    "support_tickets": (
        "support_tickets",
        ["client_id", "ticket_id", "customer_id", "ticket_type", "priority",
         "status", "channel", "opened_date", "resolved_date", "resolution_time_hrs",
         "ticket_text", "source"],
    ),
    # One row per login (engagement event log) — feeds point-in-time login
    # features in the temporal model. Same upload path as orders/tickets.
    "login_event": (
        "login_events",
        ["client_id", "login_id", "customer_id", "login_at", "login_channel"],
    ),
}

# ── Required-column rules per master type ────────────────────────────────────
# Columns that MUST carry at least one populated value in the uploaded file.
# `client_id` is intentionally omitted because the upload pipeline injects it
# at staging time — it's never present in the customer's Excel/CSV.
#
# This is the source of truth for the "file has rows but every required
# field is blank" guard. Keeping it here (not importing validation_router's
# VALIDATION_CONFIG) avoids a cross-module cycle and lets the two checkers
# evolve independently.
REQUIRED_COLS_PER_MASTER = {
    # last_login_date added 2026-04-24 — required because it drives the
    # login-aware churn label. Uploading a customer master with this
    # column entirely blank gets rejected at staging time, matching
    # validation_router's required_cols list.
    "customer":         ["customer_id", "customer_email", "customer_name", "account_created_date", "last_login_date"],
    "order":            ["order_id", "customer_id", "order_date", "order_value_usd"],
    "line_items":       ["line_item_id", "order_id", "customer_id", "product_id", "quantity", "unit_price_usd"],
    "product":          ["product_id", "product_name", "category_id"],
    "price":            ["price_id", "product_id", "unit_price_usd"],
    "vendor_map":       ["pv_id", "product_id", "vendor_id"],
    "category":         ["category_id", "category_name"],
    "sub_category":     ["sub_category_id", "sub_category_name", "category_id"],
    "sub_sub_category": ["sub_sub_category_id", "sub_sub_category_name", "sub_category_id"],
    "brand":            ["brand_id", "brand_name"],
    "vendor":           ["vendor_id", "vendor_name"],
    "customer_reviews": ["review_id", "customer_id", "product_id", "rating"],
    "support_tickets":  ["ticket_id", "customer_id", "ticket_type"],
    "login_event":      ["login_id", "customer_id", "login_at"],
}


def _assert_required_columns_present(df: pd.DataFrame, master_type: str, filename: str) -> None:
    """Reject files where any REQUIRED column is missing from the header row.

    Why this exists separately from _assert_file_has_real_data:
        Previously the only schema check between upload and staging was the
        "all matching columns are blank" guard, which bails only when EVERY
        required column happens to be present-but-empty. A file missing a
        single required column (e.g. customer master without last_login_date)
        passed this guard and then silently landed NULL in the DB, leaking
        through to the model. (Reported 2026-04-24 — last_login_date became
        required for the login-aware churn rule but uploads without it kept
        succeeding.)

        This new check runs BEFORE the blank-data guard and rejects with a
        specific, actionable message naming the missing column(s).

    Normalization matches _load_df_to_staging's column-name handling
    (strip + lower + space-to-underscore) so a user's "Last Login Date"
    column header is correctly recognized as last_login_date.
    """
    required = REQUIRED_COLS_PER_MASTER.get(master_type, [])
    if not required:
        return  # master type has no required cols; nothing to assert

    file_cols_normalized = {
        str(c).strip().lower().replace(" ", "_") for c in df.columns
    }
    missing = [c for c in required if c not in file_cols_normalized]

    if missing:
        # 2026-04-27: keep the detailed missing-column list in the SERVER
        # LOG (so we can debug from the backend console), but show the
        # SAME generic message to the user as the "wrong file" path
        # below. Two different rejection messages were confusing — users
        # didn't know whether they uploaded the wrong file vs the right
        # file with bad columns. Both cases mean the same thing from
        # their perspective: "this file doesn't match what we asked for."
        log.warning(
            "Upload rejected for '%s' — missing required column(s): %s | "
            "File had: %s",
            master_type, missing, sorted(file_cols_normalized),
        )
        raise HTTPException(
            status_code=400,
            detail="Wrong file uploaded. Please check you uploaded the correct file.",
        )


def _assert_file_has_real_data(df: pd.DataFrame, master_type: str, filename: str) -> None:
    """Reject files that have rows but every required column is 100% blank.

    `df.empty` already covers the "zero rows" case upstream; this catches the
    second-degree empty file — e.g. a template Excel where every data row is
    just placeholders or whitespace. Treats NaN AND empty / whitespace-only
    strings as blank.

    Assumes _assert_required_columns_present has already run, so any required
    column is guaranteed to be in df.columns by this point.
    """
    required = REQUIRED_COLS_PER_MASTER.get(master_type, [])
    if not required:
        return  # master type has no required cols; nothing to assert

    # Build the list of required columns that are actually present in the
    # uploaded file. After _assert_required_columns_present runs upstream,
    # this list equals `required`; the comprehension is kept defensively in
    # case a future caller skips the presence check.
    present = [c for c in required if c in df.columns]
    if not present:
        return  # schema mismatch is a different error class

    for col in present:
        series = df[col]
        # .notna() catches NaN / NaT; the string cast + strip filter catches
        # literal "" and whitespace-only cells that pandas parsed as strings.
        has_value = series.notna() & (series.astype(str).str.strip() != "")
        if has_value.any():
            return  # at least one real value found — file is fine

    raise HTTPException(
        status_code=400,
        detail=(
            f"{filename} has rows but every required column is blank "
            f"({', '.join(present)}). Please upload a file with actual data."
        ),
    )


# ── Sample CSV templates (downloadable from each upload tile) ────────────────
# Clients don't know our exact column names / order, so each upload module
# offers a "Download sample CSV" link. The template is generated FROM
# MASTER_TYPE_TO_TABLE — the same dict the validator reads — so the sample can
# never drift from what actually passes validation (single source of truth).
# client_id is intentionally omitted: it's auto-injected from the auth token on
# upload, and _read_file_to_df explicitly excludes it from column matching.

# *_id columns that are VARCHAR in the schema (human-readable codes). Every
# OTHER *_id column is an INTEGER PK/FK (product_id, category_id, brand_id, …)
# and must get a numeric sample value, or the template can't be uploaded into
# its integer column. (client_id is excluded from samples entirely.)
_STRING_ID_COLS = {"customer_id", "order_id", "line_item_id", "review_id", "ticket_id"}

# id column → the master file its value must line up with. Used only to enrich
# the template's type legend ("must match your X file"); a referential mismatch
# here is the single most common cause of the FK "couldn't save your data" error.
_FK_PARENT_FILE = {
    "customer_id":         "Customer",
    "order_id":            "Order",
    "product_id":          "Product",
    "category_id":         "Category",
    "sub_category_id":     "Sub-Category",
    "sub_sub_category_id": "Sub-Sub-Category",
    "brand_id":            "Brand",
    "vendor_id":           "Vendor",
    "product_price_id":    "Product Price",
}


def _sample_value(col: str, i: int) -> str:
    """A realistic placeholder for one column on example row `i` (0-based).

    Picked by column-name heuristics so the template shows clients the expected
    FORMAT — ISO dates, money with 2dp, boolean style, id style — not just bare
    headers. Every column gets a non-blank value so the template itself passes
    the `_assert_file_has_real_data` guard.
    """
    c = col.lower()
    n = i + 1
    # active / not_available are SMALLINT flags in the schema → 1/0 (NOT "true",
    # which fails to load into a smallint column).
    if c in ("active", "not_available"):
        return "1" if i % 2 == 0 else "0"
    # email_opt_in / sms_opt_in are real BOOLEAN columns → true/false.
    if c in ("email_opt_in", "sms_opt_in"):
        return "true" if i % 2 == 0 else "false"
    if c.endswith("email"):
        return f"customer{n}@example.com"
    if c.endswith("_date"):                       # ISO 8601, per the upload notice
        return f"2026-{(i % 9) + 1:02d}-15"
    if c.endswith("_usd"):   # every money column ends in _usd; "price_id" is an id, not money
        return f"{n * 10 + 9.99:.2f}"
    if "phone" in c or "contact_no" in c:
        return f"+1-555-01{n:02d}"
    if c == "rating":
        return str((i % 5) + 1)
    if c in ("quantity", "qty_min", "qty_max", "order_item_count", "resolution_time_hrs"):
        return str(n * 5)
    if c in ("order_status", "item_status", "status"):
        return ("completed", "pending", "shipped")[i % 3]
    if c == "sentiment":
        return ("positive", "neutral", "negative")[i % 3]
    if c == "priority":
        return ("high", "medium", "low")[i % 3]
    if c == "payment_method":
        return ("credit_card", "paypal", "debit_card")[i % 3]
    if c in ("channel", "registration_channel"):
        return ("web", "mobile", "store")[i % 3]
    if c == "preferred_device":
        return ("mobile", "desktop", "tablet")[i % 3]
    if c == "ticket_type":
        return ("billing", "delivery", "product")[i % 3]
    if c == "country_code":
        return "US"
    if c == "state":
        return ("TX", "CA", "NY")[i % 3]
    if c == "city":
        return ("Dallas", "San Jose", "New York")[i % 3]
    if c == "zip_code":
        return f"7{n:04d}"
    if c.endswith("address"):
        return f"{100 + n} Main St"
    if c == "coupon_code":
        return f"SAVE{n * 5}"
    if c == "sku":
        return f"SKU-{n:05d}"
    if c == "qty_range_label":
        return ("1-10", "11-50", "51-100")[i % 3]
    if c == "category_hint":
        return ("Electronics", "Apparel", "Home")[i % 3]
    if c == "source":
        return ("email", "phone", "chat", "portal")[i % 4]
    if c.endswith("_text"):
        return ("Great product!", "Works as expected.", "Would buy again.")[i % 3]
    if c.endswith("description"):
        return "Sample description"
    if c == "customer_name":
        return ("John Doe", "Jane Smith", "Sam Lee")[i % 3]
    if c == "product_name":
        return ("Wireless Mouse", "USB-C Cable", "Laptop Stand")[i % 3]
    if c == "brand_name":
        return ("Acme", "Globex", "Initech")[i % 3]
    if c == "vendor_name":
        return ("Acme Supply Co", "Globex Distribution", "Initech Traders")[i % 3]
    if c == "category_name":
        return ("Electronics", "Apparel", "Home & Kitchen")[i % 3]
    if c == "sub_category_name":
        return ("Accessories", "Cables", "Stands")[i % 3]
    if c == "sub_sub_category_name":
        return ("Mice", "USB-C", "Laptop")[i % 3]
    if c.endswith("_id"):
        if c in _STRING_ID_COLS:                   # varchar code columns
            base = c[:-3]
            prefix = {
                "customer": "CUST", "order": "ORD", "line_item": "LI",
                "review": "REV", "ticket": "TKT",
            }.get(base) or (base.upper().replace("_", "")[:4] or "ID")
            return f"{prefix}-{n:05d}"
        # Every other *_id is an INTEGER PK/FK in the schema. Emit a plain,
        # row-aligned integer: it fits the INTEGER column AND the same row index
        # yields the same id across files, so the full template set stays
        # referentially consistent (line_items.product_id ↔ products.product_id,
        # products.category_id ↔ categories.category_id, …).
        return str(n)
    return f"sample_{c}"


def _column_type_hint(col: str, show_fk: bool = True) -> str:
    """Human-readable expected type/format for a column, shown in the sample
    template's legend so clients know what each column should contain. FK
    columns also name the file they must line up with — the mismatch that
    triggers the FK 'couldn't save your data' error on commit."""
    c = col.lower()
    if c in ("active", "not_available"):
        base = "1 (yes) or 0 (no)"
    elif c in ("email_opt_in", "sms_opt_in"):
        base = "true or false"
    elif c.endswith("_date"):
        base = "date (YYYY-MM-DD)"
    elif c.endswith("_usd"):
        base = "amount with 2 decimals (e.g. 19.99)"
    elif c == "rating":
        base = "whole number 1-5"
    elif c in ("quantity", "qty_min", "qty_max", "order_item_count", "resolution_time_hrs"):
        base = "whole number"
    elif c.endswith("email"):
        base = "email address"
    elif "phone" in c or "contact_no" in c:
        base = "phone number"
    elif c in _STRING_ID_COLS:
        base = "text code (see example row)"
    elif c.endswith("_id"):
        base = "whole number"
    else:
        base = "text"
    if show_fk:
        parent = _FK_PARENT_FILE.get(c)
        if parent:
            base += f" - must match a row in your {parent} file"
    return base


def _sample_headers_and_rows(master_type: str, n_rows: int = 2):
    """(headers, example_rows, pk) for a master type — shared by the CSV and
    XLSX template builders. Headers come from MASTER_TYPE_TO_TABLE in canonical
    order, minus client_id (auto-injected on upload). pk is the entity's own id
    (the first column), which is never a foreign key."""
    if master_type not in MASTER_TYPE_TO_TABLE:
        raise ValueError(f"Unknown master type: {master_type}")
    _, expected_cols = MASTER_TYPE_TO_TABLE[master_type]
    headers = [c for c in expected_cols if c != "client_id"]
    rows = [[_sample_value(c, i) for c in headers] for i in range(n_rows)]
    pk = headers[0] if headers else None
    return headers, rows, pk


def build_sample_csv(master_type: str, n_rows: int = 2) -> str:
    """Clean sample CSV: header + a couple of example rows, nothing else.

    Per-column type/format guidance lives on the XLSX template's Instructions
    sheet (build_sample_xlsx), so the CSV stays a plain, uploadable grid.
    """
    headers, rows, _ = _sample_headers_and_rows(master_type, n_rows)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(r)
    return buf.getvalue()


def _xlsx_cell(col: str, val: str):
    """Coerce a string sample value to its native type so the XLSX cell isn't a
    'number stored as text' warning — mirrors the column's schema type. Codes,
    dates, booleans and free text stay as strings."""
    c = col.lower()
    try:
        if c.endswith("_usd"):
            return float(val)
        if c in ("rating", "active", "not_available", "quantity", "qty_min",
                 "qty_max", "order_item_count", "resolution_time_hrs"):
            return int(val)
        if c.endswith("_id") and c not in _STRING_ID_COLS:
            return int(val)
    except (TypeError, ValueError):
        return val
    return val


def build_sample_xlsx(master_type: str, n_rows: int = 2) -> bytes:
    """User-friendly .xlsx template with two sheets:

      • 'Template'     — styled header + example rows the client edits/uploads.
      • 'Instructions' — a per-column guide (expected format, and which file each
                         id must match). Only the FIRST sheet is read on upload,
                         so the guide never interferes with parsing.

    Being a real .xlsx (not a CSV) means no encoding/mojibake issues and a clean
    spreadsheet view — no '#' comment rows.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    headers, rows, pk = _sample_headers_and_rows(master_type, n_rows)
    friendly = _TABLE_FRIENDLY.get(MASTER_TYPE_TO_TABLE[master_type][0], master_type)

    wb = Workbook()

    # ── Sheet 1: Template — the grid the client fills in / the upload reads ──
    ws = wb.active
    ws.title = "Template"
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F5496")
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([_xlsx_cell(c, v) for c, v in zip(headers, r)])
    ws.freeze_panes = "A2"  # keep headers visible while scrolling
    for i, c in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(c) + 3)

    # ── Sheet 2: Instructions — per-column guide; ignored on upload ──
    info = wb.create_sheet("Instructions")
    info["A1"] = f"{friendly} file - how to fill it in"
    info["A1"].font = Font(bold=True, size=13)
    info["A3"] = "1. Open the 'Template' tab and replace the example rows with your own data."
    info["A4"] = "2. Keep the column headers exactly as they are. Do not add a client id column."
    info["A5"] = "3. Save as .xlsx or .csv and upload. This Instructions tab is ignored on upload."
    info["A7"], info["B7"] = "Column", "Expected format"
    info["A7"].font = info["B7"].font = Font(bold=True)
    for n, c in enumerate(headers, start=8):
        info[f"A{n}"] = c
        info[f"B{n}"] = _column_type_hint(c, show_fk=(c != pk))
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 64

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# Tables that affect the materialized view — refresh after commit if these were loaded
MV_TRIGGER_TABLES = {"customers", "orders", "line_items", "products",
                     "categories", "brands", "product_prices",
                     "customer_reviews", "support_tickets"}


# ── Conflict behavior per master type ────────────────────────────────────────
# When committing staging → real, a row's primary key may already exist (user
# re-uploading the same file, or a prior commit that was partially re-applied).
# Without explicit ON CONFLICT handling every such row raises UniqueViolation
# and rolls back the entire batch — which is brittle and makes retries
# impossible.
#
# Strategy:
#   - Catalog / dimension tables → UPSERT ('update'): re-uploads refresh the
#     non-PK columns, so corrected data overwrites old.
#   - Transactional / fact tables → SKIP ('nothing'): we never silently
#     overwrite a historical order, line_item, review, or ticket. A duplicate
#     PK is simply ignored and the rest of the batch goes through.
#
# Format: master_type → (pk_columns_tuple, action)
COMMIT_CONFLICT = {
    # Dimension / catalog (UPSERT — safe to refresh)
    "category":         (("client_id", "category_id"),         "update"),
    "sub_category":     (("client_id", "sub_category_id"),     "update"),
    "sub_sub_category": (("client_id", "sub_sub_category_id"), "update"),
    "brand":            (("client_id", "brand_id"),            "update"),
    "vendor":           (("client_id", "vendor_id"),           "update"),
    "product":          (("client_id", "product_id"),          "update"),
    "price":            (("client_id", "price_id"),            "update"),
    "vendor_map":       (("client_id", "pv_id"),               "update"),
    "customer":         (("client_id", "customer_id"),         "update"),
    # Fact / transactional (DO NOTHING — don't overwrite history)
    "order":            (("client_id", "order_id"),            "nothing"),
    "line_items":       (("client_id", "line_item_id"),        "nothing"),
    "customer_reviews": (("client_id", "review_id"),           "nothing"),
    "support_tickets":  (("client_id", "ticket_id"),           "nothing"),
    "login_event":      (("client_id", "login_id"),            "nothing"),
}


# ── Commit order ─────────────────────────────────────────────────────────────
# When committing a batch, we INSERT from staging → real in this order.
# With DEFERRABLE FKs, order doesn't affect correctness (all checks happen at
# COMMIT), but parents-before-children is still more efficient.
COMMIT_ORDER = [
    "category", "vendor",
    "sub_category", "brand",
    "sub_sub_category",
    "product",
    "price", "vendor_map",
    "customer",
    "order",
    "line_items",
    "customer_reviews", "support_tickets", "login_event",
]


# ── Pre-flight FK validation rules ───────────────────────────────────────────
# Each tuple: (staging_table, staging_col, parent_real_table, parent_real_col,
#              parent_staging_table, fk_name)
# For every staging row where staging_col IS NOT NULL, the referenced key
# must exist EITHER in the parent's real table OR in the parent's staging
# table for the same client_id and batch_id.
CATALOG_FK_CHECKS = [
    ("staging_sub_categories",          "category_id",         "categories",          "category_id",         "staging_categories",         "sub_categories_category_fk"),
    ("staging_sub_sub_categories",      "sub_category_id",     "sub_categories",      "sub_category_id",     "staging_sub_categories",     "sub_sub_categories_sub_category_fk"),
    ("staging_sub_sub_categories",      "category_id",         "categories",          "category_id",         "staging_categories",         "sub_sub_categories_category_fk"),
    ("staging_brands",                  "vendor_id",           "vendors",             "vendor_id",           "staging_vendors",            "brands_vendor_fk"),
    ("staging_products",                "category_id",         "categories",          "category_id",         "staging_categories",         "products_category_fk"),
    ("staging_products",                "sub_category_id",     "sub_categories",      "sub_category_id",     "staging_sub_categories",     "products_sub_category_fk"),
    ("staging_products",                "sub_sub_category_id", "sub_sub_categories",  "sub_sub_category_id", "staging_sub_sub_categories", "products_sub_sub_category_fk"),
    ("staging_products",                "brand_id",            "brands",              "brand_id",            "staging_brands",             "products_brand_fk"),
    ("staging_products",                "product_price_id",    "product_prices",      "price_id",            "staging_product_prices",     "fk_products_price"),
    ("staging_product_prices",          "product_id",          "products",            "product_id",          "staging_products",           "product_prices_product_fk"),
    ("staging_product_vendor_mapping",  "product_id",          "products",            "product_id",          "staging_products",           "pvm_product_fk"),
    ("staging_product_vendor_mapping",  "brand_id",            "brands",              "brand_id",            "staging_brands",             "pvm_brand_fk"),
    ("staging_product_vendor_mapping",  "vendor_id",           "vendors",             "vendor_id",           "staging_vendors",            "pvm_vendor_fk"),
    ("staging_line_items",              "product_id",          "products",            "product_id",          "staging_products",           "line_items_product_fk"),
    ("staging_customer_reviews",        "product_id",          "products",            "product_id",          "staging_products",           "customer_reviews_product_fk"),
]


# ── Batch tracking helpers ────────────────────────────────────────────────────
def _get_pending_batch_id(conn, client_id: str) -> Optional[str]:
    """Return UUID string of the client's pending batch, or None if none exists."""
    result = conn.execute(
        text("SELECT batch_id::text FROM upload_batches "
             "WHERE client_id = :cid AND status = 'pending'"),
        {"cid": client_id},
    ).fetchone()
    return result[0] if result else None


def _get_or_create_pending_batch(conn, client_id: str) -> str:
    """Return pending batch UUID, creating one if none exists."""
    existing = _get_pending_batch_id(conn, client_id)
    if existing:
        return existing

    new_id = str(uuid.uuid4())
    conn.execute(
        text("INSERT INTO upload_batches (batch_id, client_id) VALUES (:bid, :cid)"),
        {"bid": new_id, "cid": client_id},
    )
    log.info("Created new pending batch %s for client %s", new_id, client_id)
    return new_id


def _staging_table_for(master_type: str) -> str:
    """Return the staging table name for a given master type (e.g. 'customer' → 'staging_customers')."""
    real_table, _ = MASTER_TYPE_TO_TABLE[master_type]
    return f"staging_{real_table}"


def _count_batch_staging_rows(conn, batch_id: str, client_id: str) -> int:
    """
    Return the total number of staging rows for a batch across ALL staging
    tables. Used to detect orphaned pending batches (batch_id exists but every
    staging table is empty), which can happen when a user uploads a file then
    hits Remove — the staging rows go away but the upload_batches row stays
    with status='pending'.
    """
    total = 0
    for master_type in COMMIT_ORDER:
        real_table, _ = MASTER_TYPE_TO_TABLE[master_type]
        staging_table = f"staging_{real_table}"
        row = conn.execute(
            text(f"SELECT COUNT(*) FROM {staging_table} "
                 f"WHERE batch_id = :bid AND client_id = :cid"),
            {"bid": batch_id, "cid": client_id},
        ).fetchone()
        total += (row[0] if row else 0)
    return total


def _cleanup_empty_batch(conn, batch_id: str, client_id: str) -> bool:
    """
    If `batch_id` has zero staging rows across every staging table, delete
    the upload_batches row so a fresh upload starts with a fresh batch.
    Returns True if the batch was cleaned up, False if it still has rows.
    """
    remaining = _count_batch_staging_rows(conn, batch_id, client_id)
    if remaining == 0:
        conn.execute(
            text("DELETE FROM upload_batches WHERE batch_id = :bid AND status = 'pending'"),
            {"bid": batch_id},
        )
        log.info("Cleaned up empty pending batch %s for client %s", batch_id, client_id)
        return True
    return False


def _humanize_staging_error(exc: Exception, master_type: str = "") -> str:
    """Translate a raw DB staging exception into a SHORT, user-friendly message.

    A staging INSERT failure used to surface the full SQLAlchemy exception — the
    SQL statement plus every bound parameter set — straight into the UI, which is
    intimidating and not actionable. The caller already logs the full error
    (exc_info=True); here we return only a friendly, actionable sentence and NEVER
    leak the SQL or parameters. Covers the common Postgres staging failures and
    falls back to a generic "wrong file" message.
    """
    import re

    orig = getattr(exc, "orig", None) or exc            # unwrap SQLAlchemy → psycopg2
    name = type(orig).__name__
    pgcode = getattr(orig, "pgcode", "") or ""
    detail = str(orig)
    m = re.search(r'column "([^"]+)"', detail)
    col = f" (column '{m.group(1)}')" if m else ""
    label = f" for {master_type.replace('_', ' ').title()}" if master_type else ""

    if name in ("DatatypeMismatch", "InvalidTextRepresentation") or pgcode in ("42804", "22P02"):
        return (f"Wrong file or wrong format{label}{col}: a value doesn't match the expected "
                "type — for example text or true/false where a number is expected. Please "
                "check you uploaded the correct file for this slot.")
    if name == "UniqueViolation" or pgcode == "23505":
        return "Some rows were already uploaded (duplicate IDs); the existing data was kept."
    if name == "ForeignKeyViolation" or pgcode == "23503":
        return ("This file references IDs that don't exist yet — upload the parent files first "
                "(for example categories, brands and vendors before products).")
    if name == "NotNullViolation" or pgcode == "23502":
        return f"A required column{col} is empty in one or more rows. Please fill it in and retry."
    if name in ("NumericValueOutOfRange", "StringDataRightTruncation") or pgcode in ("22003", "22001"):
        return f"A value{col} is too large or too long for its column. Please check the file."
    if name == "UndefinedColumn" or pgcode == "42703":
        return ("The file has a column that doesn't belong in this slot. Please check you "
                "uploaded the correct file.")
    return ("Wrong file uploaded, or the file has an unexpected format. Please check you "
            "uploaded the correct file for this slot.")


def _build_sync_connector(provider: str, client_id: str):
    """Build the tenant's connector for a live 'Sync into batch' pull. Raises a
    clear HTTPException when the provider is unknown or not yet configured."""
    from ml.connectors.registry import CONNECTOR_REGISTRY
    cls = CONNECTOR_REGISTRY.get(provider)
    if cls is None:
        raise HTTPException(status_code=400, detail=f"Unknown provider '{provider}'.")
    try:
        connector = cls.from_client(engine, client_id, require_enabled=False)
    except Exception as exc:  # noqa: BLE001 — crypto/config errors → friendly message
        log.warning("sync connector build failed %s/%s: %s", client_id, provider, exc)
        raise HTTPException(status_code=500,
                            detail=f"Could not build the {provider} connector — check its settings.")
    if connector is None:
        raise HTTPException(status_code=400,
                            detail=f"No {provider} credentials saved — configure the integration first.")
    return connector


def _records_to_df(records, signal_kind: str):
    """Convert connector RawTicket/RawReview records to the column shape the
    ticket/review staging path expects (client_id is injected at staging time)."""
    rows = []
    for r in records:
        if signal_kind == "ticket":
            rows.append({
                "ticket_id": r.ticket_id, "customer_id": r.customer_id,
                "ticket_type": r.ticket_type, "priority": r.priority, "status": r.status,
                "opened_date": r.opened_date, "resolved_date": r.resolved_date,
                "ticket_text": r.text, "source": r.source,
            })
        else:  # review
            rows.append({
                "review_id": r.review_id, "customer_id": r.customer_id,
                "rating": r.rating, "review_text": r.text,
                "review_date": r.review_date, "source": r.source,
            })
    return pd.DataFrame(rows)


def _resolve_customers_in_df(df, client_id: str):
    """For ticket/review uploads: rewrite each row's customer_id to a known
    customer (id→email). Drops unmatched rows. Returns (matched_df, report).

    "Known" = committed customers OR customers staged in the client's CURRENT
    pending batch — so a brand-new client can onboard customers + tickets in one
    batch (the deferred-FK commit inserts the customers first). Tickets for a
    customer that is neither committed nor staged are still dropped."""
    from ml.connectors.sources import resolve_customer_id

    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT customer_id, LOWER(customer_email) FROM customers WHERE client_id = :c
                UNION
                SELECT customer_id, LOWER(customer_email) FROM staging_customers
                 WHERE client_id = :c
                   AND batch_id = (SELECT batch_id FROM upload_batches
                                   WHERE client_id = :c AND status = 'pending')
            """),
            {"c": client_id},
        ).fetchall()
    by_id = {r[0] for r in rows}
    by_email = {r[1]: r[0] for r in rows if r[1]}

    matched, skipped_sample = [], []
    for idx, row in df.iterrows():
        rid = resolve_customer_id(row.to_dict(), by_id, by_email)
        if rid is not None:
            matched.append((idx, rid))
        elif len(skipped_sample) < 10:
            ref = row.get("customer_id")
            if ref is None or str(ref).strip() == "":
                ref = row.get("customer_email") or row.get("email") or ""
            skipped_sample.append(str(ref))

    if matched:
        idxs, rids = zip(*matched)
        matched_df = df.loc[list(idxs)].copy()
        matched_df["customer_id"] = list(rids)
    else:
        matched_df = df.iloc[0:0].copy()

    report = {
        "matched": len(matched),
        "skipped": int(len(df) - len(matched)),
        "skippedSample": skipped_sample,
    }
    return matched_df, report


def _load_df_to_staging(df: pd.DataFrame, master_type: str, client_id: str) -> dict:
    """
    Insert a pandas DataFrame into the STAGING table for this master type.

    REPLACE SEMANTICS: if the same master_type was already uploaded in the
    current pending batch, those staging rows are deleted first. Last
    upload wins — so the user can fix a typo and re-upload.

    The real tables are NOT touched here. Data moves from staging to real
    only when the user calls POST /uploads/commit.

    Returns: {staged, batch_id, staging_table, rows_staged} on success,
             {staged: False, reason} on failure.
    """
    if master_type not in MASTER_TYPE_TO_TABLE:
        return {"staged": False, "reason": f"No table mapping for {master_type}"}

    real_table, expected_cols = MASTER_TYPE_TO_TABLE[master_type]
    staging_table = f"staging_{real_table}"

    # ── Normalize column names in the uploaded DataFrame ────────────────
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    # Always stamp client_id from the authenticated form field — NEVER trust
    # whatever client_id the CSV was saved with.
    #
    # Why: this is a multi-tenant system. A client admin (say, Sams Club) might
    # re-use a CSV that was originally exported with client_id='CLT-001'
    # (Walmart) baked into every row. The old code only injected client_id if
    # the column was missing; if it was present with CLT-001, the rows went to
    # staging with client_id='CLT-001' but batch_id for CLT-004. The downstream
    # DELETE / preflight / INSERT all filter by WHERE client_id = :cid so
    # nothing actually got committed — the upload silently did nothing.
    #
    # The fix: if the expected schema has client_id at all, overwrite whatever
    # the file had. The form/token is the source of truth for tenancy.
    if "client_id" in expected_cols:
        if "client_id" in df.columns:
            mismatched = (df["client_id"].astype(str) != str(client_id)).sum()
            if mismatched:
                log.warning(
                    "Overriding client_id in %d/%d uploaded rows "
                    "(file had a different client_id than target %s)",
                    int(mismatched), len(df), client_id,
                )
            df["client_id"] = client_id
        else:
            df.insert(0, "client_id", client_id)

    # Keep only columns that exist in both the DataFrame and the expected list
    available_cols = [c for c in expected_cols if c in df.columns]
    log.info(
        "Staging prep: %s → %s | expected %d cols, found %d matching | df cols: %s",
        master_type, staging_table, len(expected_cols), len(available_cols), list(df.columns)[:5]
    )
    if not available_cols:
        return {
            "staged": False,
            "reason": f"No matching columns. Expected: {expected_cols}, got: {list(df.columns)}",
        }

    df_to_load = df[available_cols].copy()

    # ── Drop rows where the business PK is null ─────────────────────────
    # Use the first non-client_id column as the PK for null checking.
    # (Previous code used available_cols[0] which was always 'client_id'
    #  when client_id is in expected_cols — making the dropna a no-op.)
    pk_col = next((c for c in available_cols if c != "client_id"), available_cols[0])
    df_to_load = df_to_load.dropna(subset=[pk_col])

    if df_to_load.empty:
        return {"staged": False, "reason": f"All rows had null {pk_col}"}

    # ── Clean up data types before insert ─────────────────────────────
    # Boolean columns: convert NaN/float to proper True/False/None
    # This fixes: "column email_opt_in is boolean but expression is text"
    for col in df_to_load.columns:
        if df_to_load[col].dtype == 'float64':
            unique_vals = set(df_to_load[col].dropna().unique())
            if unique_vals <= {0.0, 1.0, True, False}:
                df_to_load[col] = df_to_load[col].apply(
                    lambda x: None if pd.isna(x) else bool(x)
                )

    # ── Build the INSERT statement targeting the staging table ──────────
    # Staging tables have the same columns as real tables PLUS batch_id
    # and staging_row_id (BIGSERIAL, auto-filled). We only need to supply
    # the real columns + batch_id; staging_row_id is generated by Postgres.
    cols_with_batch = available_cols + ["batch_id"]
    col_list = ", ".join(cols_with_batch)
    placeholders = ", ".join([f":{c}" for c in cols_with_batch])
    insert_sql = text(
        f"INSERT INTO {staging_table} ({col_list}) VALUES ({placeholders})"
    )

    rows_staged = 0

    try:
        # engine.begin() wraps everything in a single transaction — either
        # ALL rows land in staging, or none do (no half-uploads).
        with engine.begin() as conn:
            # 1. Get or create the client's pending batch
            batch_id = _get_or_create_pending_batch(conn, client_id)

            # 2. REPLACE: wipe any existing rows for this batch + master_type
            #    so re-uploads overwrite instead of accumulating duplicates.
            delete_result = conn.execute(
                text(f"DELETE FROM {staging_table} "
                     f"WHERE batch_id = :bid AND client_id = :cid"),
                {"bid": batch_id, "cid": client_id},
            )
            rows_replaced = delete_result.rowcount

            # 3. Convert DataFrame to records, clean NaN → None, stamp batch_id
            records = []
            for record in df_to_load.to_dict("records"):
                clean = {}
                for k, v in record.items():
                    if isinstance(v, float) and pd.isna(v):
                        clean[k] = None
                    elif v is None:
                        clean[k] = None
                    else:
                        clean[k] = v
                clean["batch_id"] = batch_id
                records.append(clean)

            # 4. Insert in chunks of 500 rows to keep memory/round-trips reasonable
            batch_size = 500
            for i in range(0, len(records), batch_size):
                chunk = records[i : i + batch_size]
                result = conn.execute(insert_sql, chunk)
                rows_staged += result.rowcount

        log.info(
            "Staging load OK: %s → %s | batch=%s | %d rows staged (replaced %d)",
            master_type, staging_table, batch_id, rows_staged, rows_replaced,
        )

        return {
            "staged": True,
            "staging_table": staging_table,
            "batch_id": batch_id,
            "rows_staged": rows_staged,
            "rows_replaced": rows_replaced,
        }

    except Exception as e:
        # Full technical error (SQL + bound params) goes to the SERVER LOG only.
        log.error("Staging load FAILED for %s: %s", staging_table, e, exc_info=True)
        # The UI gets a short, actionable message — never the raw psycopg2 dump.
        return {"staged": False, "reason": _humanize_staging_error(e, master_type)}


def _is_decorative_row(df: pd.DataFrame) -> bool:
    """
    Check if a DataFrame's column headers look like a decorative/title row
    rather than actual data column names.

    A decorative row typically has:
    - Most columns named "Unnamed: X" (because they were empty in Excel)
    - At least one column with a long string (>30 chars) or containing
      pipe "|" characters (used for decorative descriptions/separators)
    """
    unnamed_count = sum(1 for c in df.columns if str(c).startswith("Unnamed"))
    has_title_col = any(
        len(str(c)) > 30 or "|" in str(c)
        for c in df.columns
    )
    return has_title_col and unnamed_count >= len(df.columns) // 2


def _read_file_to_df(file_bytes: bytes, filename: str, master_type: str = None) -> pd.DataFrame:
    """
    Read CSV or Excel file bytes into a pandas DataFrame.

    SMART HEADER DETECTION (up to 5 rows):
    ──────────────────────────────────────
    Some Excel files have decorative title/description rows before the
    actual column headers. For example:

        Row 1: "👤 Customer Master | 100 rows | ..."     ← decorative
        Row 2: "client_id", "customer_id", ...            ← real headers
        Row 3: "CLT-002", "CUST-001", ...                 ← data

    Or even TWO decorative rows (like the price master):

        Row 1: "💲 Product Price Master | ..."            ← decorative
        Row 2: "qty_min / qty_max are unit counts | ..."  ← also decorative
        Row 3: "client_id", "price_id", "product_id", ... ← real headers
        Row 4: data

    We scan up to 5 rows looking for the first non-decorative row.
    If none found within 5 rows, we reject the file with an error.

    COLUMN VALIDATION:
    After finding headers, we check if ANY of the expected DB columns
    match the file's columns. If zero match (excluding auto-injected
    client_id), the file is wrong for this master type → error.
    """
    lower = filename.lower()
    if lower.endswith(".csv"):
        # A downloaded sample template carries a leading '#' type/format legend.
        # Skip those lines so the client can upload the template as-is. Real
        # user files (no leading '#') take the original path unchanged.
        if file_bytes[:64].decode("utf-8-sig", errors="ignore").lstrip().startswith("#"):
            lines = file_bytes.decode("utf-8-sig", errors="replace").splitlines()
            k = 0
            while k < len(lines) and lines[k].lstrip().startswith("#"):
                k += 1
            df = pd.read_csv(io.StringIO("\n".join(lines[k:])))
        else:
            df = pd.read_csv(io.BytesIO(file_bytes))
    elif lower.endswith((".xlsx", ".xls")):
        # ── Scan up to 5 rows to find real column headers ──────────
        header_row = 0
        found_headers = False

        for try_header in range(5):
            try:
                test_df = pd.read_excel(io.BytesIO(file_bytes), header=try_header)
            except Exception:
                break  # ran out of rows

            if test_df.empty and try_header > 0:
                break  # no more data rows

            if _is_decorative_row(test_df):
                header_row = try_header + 1
                log.info("Row %d in %s is decorative — skipping", try_header + 1, filename)
                continue
            else:
                found_headers = True
                break

        if not found_headers and header_row >= 5:
            raise HTTPException(
                status_code=400,
                detail=f"Could not find column headers in the first 5 rows of {filename}. "
                       f"Please make sure the file has proper column headers "
                       f"(e.g., 'vendor_id', 'vendor_name', etc.).",
            )

        if header_row > 0:
            log.info("Detected title row in %s — using row %d as headers", filename, header_row + 1)

        df = pd.read_excel(io.BytesIO(file_bytes), header=header_row)
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file format: {filename}. Use .csv, .xlsx, or .xls",
        )

    # ── Column validation: check that at least some columns match ──────
    if master_type and master_type in MASTER_TYPE_TO_TABLE:
        _, expected_cols = MASTER_TYPE_TO_TABLE[master_type]
        # Normalize file columns the same way _load_df_to_staging does
        normalized_cols = [c.strip().lower().replace(" ", "_") for c in df.columns]
        # Don't count client_id — it gets auto-injected anyway
        expected_without_cid = [c for c in expected_cols if c != "client_id"]
        matching = [c for c in expected_without_cid if c in normalized_cols]

        if len(matching) == 0:
            # User-facing message stays short and friendly — the wall of
            # column names that used to appear here was intimidating and
            # not actionable for the typical client user. Detailed
            # expected-vs-found is logged server-side so devs can still
            # diagnose mis-uploads from the logs without exposing internal
            # column names in the UI. (2026-04-24)
            log.warning(
                "Wrong-file upload rejected for '%s'. Expected: %s | Found: %s",
                master_type, expected_without_cid, normalized_cols,
            )
            raise HTTPException(
                status_code=400,
                detail="Wrong file uploaded. Please check you uploaded the correct file.",
            )
        log.info("Column validation for %s: %d/%d expected columns matched",
                 master_type, len(matching), len(expected_without_cid))

    return df


# ── Commit / discard endpoints (declared BEFORE the /{master_type} wildcard) ──
# FastAPI matches routes in declaration order. If /uploads/{master_type} were
# declared first, a POST to /uploads/commit would be routed there with
# master_type="commit" and get rejected. Declaring literals first fixes that.

@router.post("/uploads/commit")
def commit_batch(
    request: Request,
    clientId: str = Query(...),
    user: dict = Depends(get_current_user),  # Audit fix #1
):
    """
    Validate + commit the pending batch to real tables.

    Two phases:
      1. Pre-flight FK check (reads staging + real, fails fast with details)
      2. Transactional commit with DEFERRABLE FKs (atomic — all or nothing)

    If anything fails, staging data is preserved and the user can fix and
    retry. On success, staging is cleared and the batch is marked committed.
    """
    _require_client_access(user, clientId)
    with engine.begin() as conn:
        batch_id = _get_pending_batch_id(conn, clientId)
        if not batch_id:
            raise HTTPException(
                status_code=400,
                detail="No pending batch to commit. Upload at least one file first.",
            )

        # Phase 1: FK validation
        violations = _preflight_fk_check(conn, batch_id, clientId)
        if violations:
            # Log the raw violations for diagnostics; return a plain-English,
            # actionable message to the client (the raw fk/table names are
            # meaningless to them and rendered as a JSON wall in the UI).
            log.warning("FK preflight failed: client=%s batch=%s violations=%s",
                        clientId, batch_id, violations)
            raise HTTPException(
                status_code=400,
                detail=_humanize_fk_violations(violations),
            )

        # Phase 2: deferred-FK transactional insert
        # engine.begin() above already started a transaction.
        conn.execute(text("SET CONSTRAINTS ALL DEFERRED"))
        result = _commit_batch(conn, batch_id, clientId)

    # Refresh the materialized view OUTSIDE the commit transaction — it's
    # a heavy operation and shouldn't hold the commit lock longer than needed.
    if result["mvNeedsRefresh"]:
        try:
            with engine.begin() as mv_conn:
                mv_conn.execute(text("REFRESH MATERIALIZED VIEW mv_customer_features"))
                log.info("Refreshed mv_customer_features after commit of batch %s", batch_id)
        except Exception as e:
            log.error("MV refresh after commit failed: %s", e, exc_info=True)
            # Don't fail the whole commit — MV can be refreshed manually.
            result["mvRefreshWarning"] = str(e)

    # Audit fix #4: record the successful commit so the audit log can
    # reconstruct who promoted which batch into the real tables.
    total_rows = sum(result["perTypeRows"].values())
    log_audit_event(
        request,
        action_type="batch_commit",
        details=(
            f"batch={batch_id} · {total_rows:,} rows committed across "
            f"{len([k for k, v in result['perTypeRows'].items() if v > 0])} table(s)"
        ),
        client_id=clientId,
        user_id=user.get("id"),
        user_email=user.get("email"),
        outcome="success",
    )

    return {
        "committed": True,
        "batchId": batch_id,
        "rowsCommitted": result["perTypeRows"],
        "mvRefreshed": result["mvNeedsRefresh"],
    }


def _discard_pending_batch_for_client(client_id: str):
    """Discard a client's PENDING upload batch: delete its staging rows, mark the
    batch 'discarded', and remove its on-disk files. Returns (batch_id, rows_deleted),
    or None if the client had no pending batch.

    Shared by the explicit /uploads/discard endpoint AND by logout cleanup, so an
    unsaved (uncommitted) batch never follows a user into the next session.
    """
    with engine.begin() as conn:
        batch_id = _get_pending_batch_id(conn, client_id)
        if not batch_id:
            return None

        total_deleted = 0
        for master_type in COMMIT_ORDER:
            real_table, _ = MASTER_TYPE_TO_TABLE[master_type]
            staging_table = f"staging_{real_table}"
            result = conn.execute(
                text(f"DELETE FROM {staging_table} WHERE batch_id = :bid"),
                {"bid": batch_id},
            )
            total_deleted += result.rowcount

        conn.execute(
            text("UPDATE upload_batches SET status = 'discarded', "
                 "committed_at = NOW() WHERE batch_id = :bid"),
            {"bid": batch_id},
        )

    # Best-effort: remove this client's uploaded files from disk
    try:
        for f in os.listdir(UPLOAD_DIR):
            if f.startswith(f"{client_id}_"):
                os.remove(os.path.join(UPLOAD_DIR, f))
    except Exception:
        pass

    return batch_id, total_deleted


@router.post("/uploads/discard")
def discard_batch(
    request: Request,
    clientId: str = Query(...),
    user: dict = Depends(get_current_user),  # Audit fix #1
):
    """
    Throw away the pending batch. Deletes all its staging rows and marks
    the batch 'discarded'. Uploaded files on disk are also deleted.
    """
    _require_client_access(user, clientId)
    result = _discard_pending_batch_for_client(clientId)
    if result is None:
        return {"discarded": False, "reason": "No pending batch"}
    batch_id, total_deleted = result

    # Audit fix #4: every batch discard leaves a tamper-resistant trail.
    log_audit_event(
        request,
        action_type="batch_discard",
        details=f"batch={batch_id} · {total_deleted:,} staging rows deleted",
        client_id=clientId,
        user_id=user.get("id"),
        user_email=user.get("email"),
        outcome="success",
    )

    return {
        "discarded": True,
        "batchId": batch_id,
        "rowsDeleted": total_deleted,
    }


@router.get("/uploads/sample/{master_type}")
def download_sample_template(master_type: str):
    """Download a sample .xlsx template for a master type.

    Public (no auth/tenant): the template is schema metadata only — the exact
    columns and order we expect, with example rows showing the value formats and
    an Instructions sheet describing each column. No client data is involved
    (client_id is auto-injected on upload and intentionally omitted). Declared as
    GET /uploads/sample/{type}; it does not collide with the POST/DELETE
    /uploads/{type} wildcards (different method).
    """
    if master_type not in VALID_MASTER_TYPES:
        raise HTTPException(
            status_code=404,
            detail=f"Unknown master type: {master_type}. "
                   f"Valid types: {sorted(VALID_MASTER_TYPES)}",
        )
    xlsx_bytes = build_sample_xlsx(master_type)
    filename = f"{master_type}_sample_template.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/uploads/{master_type}")
async def upload_file(
    master_type: str,
    request: Request,
    file: UploadFile = File(...),
    clientId: str = Form(None),             # NOW OPTIONAL — auto-detected from token
    masterType: str = Form(None),
    source: Optional[str] = Form(None),
    sourceName: Optional[str] = Form(None),
    authorization: Optional[str] = Header(default=None),
    user: dict = Depends(get_current_user),  # Audit fix #1 — require auth
):
    """
    Upload a CSV or Excel file for a specific master type.

    HOW client_id AUTO-DETECTION WORKS:
    ───────────────────────────────────
    1. If clientId is provided in the form → use it (backward compatible)
    2. If NOT provided → look at the logged-in user's token
       a. Get user's clientAccess list (e.g., ["CLT-002"])
       b. If user has access to exactly ONE client → use that automatically
       c. If user has access to MULTIPLE clients → require clientId in form
       d. If user is super_admin with "*" access → require clientId in form

    This means: a Costco user just uploads their Excel file.
    The system knows they're Costco (CLT-002) and tags every row automatically.
    They NEVER need to add a client_id column to their Excel sheets!
    """
    # Validate master type
    if master_type not in VALID_MASTER_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown master type: {master_type}. "
                   f"Valid types: {sorted(VALID_MASTER_TYPES)}",
        )

    # ── Auto-detect client_id if not provided ─────────────────────────
    # `user` is now provided by Depends(get_current_user) above (audit #1
    # — auth was previously skipped when clientId was supplied in the
    # form, so anyone could upload to any tenant).
    if not clientId:
        client_access = user.get("clientAccess", [])

        if "*" in client_access:
            # Super admins must specify which client they're uploading for
            raise HTTPException(
                status_code=400,
                detail="You have access to all clients. Please specify clientId in the upload form.",
            )
        elif len(client_access) == 1:
            # User has access to exactly one client → auto-detect!
            clientId = client_access[0]
            log.info("Auto-detected client_id=%s from user %s", clientId, user.get("email"))
        elif len(client_access) > 1:
            # User has access to multiple clients → they need to pick one
            raise HTTPException(
                status_code=400,
                detail=f"You have access to multiple clients: {client_access}. "
                       f"Please specify clientId in the upload form.",
            )
        else:
            raise HTTPException(status_code=403, detail="You don't have access to any clients")
    else:
        # Caller supplied a clientId — verify they actually have access
        # to it (audit #1 — previously trusted the form value blindly).
        _require_client_access(user, clientId)

    # ── File size guard (audit fix #3) ────────────────────────────────
    # Reject before reading the body into memory. Some clients send
    # Content-Length; FastAPI exposes it as file.size when set.
    declared_size = getattr(file, "size", None)
    if declared_size is not None and declared_size > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large: {declared_size:,} bytes exceeds the "
                   f"{MAX_UPLOAD_BYTES:,}-byte cap. Override with the "
                   f"CRP_MAX_UPLOAD_BYTES env var if your dataset legitimately "
                   f"needs more.",
        )

    # Read the file. Even if the client lied about Content-Length, we
    # check the actual buffered size again before staging.
    file_bytes = await file.read()
    file_size = len(file_bytes)
    if file_size > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large: {file_size:,} bytes exceeds the "
                   f"{MAX_UPLOAD_BYTES:,}-byte cap.",
        )
    # Audit fix #2: sanitize the user-supplied filename before we do
    # ANYTHING with it on disk. Path traversal via `../` would otherwise
    # let a malicious upload write outside UPLOAD_DIR.
    raw_filename = file.filename or f"{master_type}.csv"
    filename = _sanitize_filename(raw_filename)
    if filename != raw_filename:
        log.info(
            "Sanitized upload filename %r → %r (client=%s, master_type=%s)",
            raw_filename, filename, clientId, master_type,
        )

    # Validate by reading with pandas
    try:
        df = _read_file_to_df(file_bytes, filename, master_type=master_type)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {str(e)}")

    if df.empty:
        raise HTTPException(status_code=400, detail="File is empty — no rows found")

    # First-line schema check: every required column must be in the header
    # row. Catches "right file type but missing a column" — e.g. customer
    # master uploaded without last_login_date. Raises 400 with the missing
    # column name(s) if so. Must run BEFORE _assert_file_has_real_data,
    # which assumes required columns exist in df.columns.
    _assert_required_columns_present(df, master_type, filename)

    # Guard against files that have rows but every required column is blank
    # (e.g. template Excel where every cell is a placeholder). Raises 400 if so.
    _assert_file_has_real_data(df, master_type, filename)

    # Source-aware ingest (tickets/reviews only): set the source on every row —
    # KEEP a pre-labeled file's source (canonicalized, e.g. JIRA->jira) and FILL
    # blank/missing values from the dropdown selection — then resolve each row to
    # a known customer, dropping unmatched rows.
    match_report = None
    if master_type in ("support_tickets", "customer_reviews"):
        from ml.connectors.sources import normalize_source, apply_source
        fill = normalize_source(source, sourceName)
        if "source" in df.columns:
            # Blank cells parse as NaN; treat them as empty so apply_source fills
            # them from the dropdown (str(NaN) would otherwise be the text 'nan').
            existing = df["source"].where(df["source"].notna(), "")
            df["source"] = existing.apply(lambda v: apply_source(v, fill))
        else:
            df["source"] = fill
        df, match_report = _resolve_customers_in_df(df, clientId)
        if df.empty:
            raise HTTPException(
                status_code=400,
                detail="None of the uploaded rows could be matched to a customer "
                       "in this client. Check the customer_id / email columns.",
            )

    row_count = len(df)
    columns = list(df.columns)

    # Save file to disk
    # NOTE (source-aware ingest): for ticket/review uploads the file written
    # here is the RAW upload — all rows, including those that were skipped
    # because their customer_id/email didn't resolve. This is intentional:
    # the on-disk file serves as a skip audit trail so ops can review
    # which rows were dropped and why. `rowCount` and the staging table
    # reflect ONLY the resolved/matched rows that were actually staged.
    try:
        safe_filename = f"{clientId}_{master_type}_{filename}"
        save_path = os.path.join(UPLOAD_DIR, safe_filename)
        with open(save_path, "wb") as f:
            f.write(file_bytes)
        log.info(f"Saved {safe_filename} ({row_count} rows, {len(columns)} columns)")
    except Exception as e:
        log.warning(f"Could not save file to disk: {e}")

    # ── LOAD into staging (NOT real tables yet) ───────────────────────
    # The rows land in staging_<table> with a batch_id. They move into the
    # real tables only when the user calls POST /uploads/commit.
    db_result = _load_df_to_staging(df, master_type, clientId)

    if not db_result.get("staged"):
        # Staging failed (bad columns, DB error, etc.). Surface the reason.
        raise HTTPException(
            status_code=400,
            detail=f"Could not stage file: {db_result.get('reason', 'unknown error')}",
        )

    # ── Audit the successful upload ─────────────────────────────────
    # The UI's "File Upload" row in the audit log comes from here.
    # `user` is the authenticated identity from Depends — no need to
    # re-resolve from the token (the previous code did, but now that
    # auth is required up front the user is always present).
    log_audit_event(
        request,
        action_type="file_upload",
        details=f"{master_type} · {filename} · {row_count:,} rows",
        client_id=clientId,
        user_id=user.get("id"),
        user_email=user.get("email"),
        outcome="success",
    )

    result = {
        "masterType": master_type,
        "fileName": filename,
        "rowCount": row_count,
        "columns": columns,
        "uploadedAt": datetime.now().isoformat(),
        "status": "staged",
        "batchId": db_result["batch_id"],
        "stagingTable": db_result["staging_table"],
        "rowsStaged": db_result["rows_staged"],
        "rowsReplaced": db_result["rows_replaced"],
    }
    if match_report is not None:
        result["matchReport"] = match_report
    return result


@router.post("/uploads/sync/{provider}")
def sync_provider_to_staging(
    provider: str,
    request: Request,
    clientId: str = Query(...),
    user: dict = Depends(get_current_user),
):
    """'Sync into batch' — pull the tenant's live records via the connector and
    route them through the SAME source-stamp + customer-match + staging path a
    CSV upload uses. Unregistered customers are dropped centrally (never an FK
    break), and the matched rows land in the pending batch for review + Save."""
    _require_client_access(user, clientId)
    connector = _build_sync_connector(provider, clientId)
    kind = getattr(connector, "signal_kind", "ticket")
    master_type = "support_tickets" if kind == "ticket" else "customer_reviews"

    try:
        records = list(connector.fetch(clientId))
    except Exception as exc:  # noqa: BLE001 — never echo the upstream body
        log.warning("sync fetch failed %s/%s: %s", clientId, provider, exc)
        raise HTTPException(status_code=502,
                            detail=f"Could not sync from {provider} — check the integration settings.")

    df = _records_to_df(records, kind)
    if df.empty:
        return {"provider": provider, "masterType": master_type, "fetched": 0,
                "matchReport": {"matched": 0, "skipped": 0, "skippedSample": []},
                "rowsStaged": 0, "batchId": None}

    fetched = len(df)
    df, match_report = _resolve_customers_in_df(df, clientId)
    if df.empty:
        return {"provider": provider, "masterType": master_type, "fetched": fetched,
                "matchReport": match_report, "rowsStaged": 0, "batchId": None}

    db_result = _load_df_to_staging(df, master_type, clientId)
    log_audit_event(
        request, action_type="file_upload",
        details=(f"sync {provider} · {master_type} · "
                 f"{match_report['matched']} matched / {match_report['skipped']} skipped"),
        client_id=clientId, user_id=user.get("id"), user_email=user.get("email"),
        outcome="success",
    )
    return {
        "provider": provider, "masterType": master_type, "fetched": fetched,
        "matchReport": match_report, "rowsStaged": db_result["rows_staged"],
        "batchId": db_result["batch_id"],
    }


@router.get("/uploads/sources")
def list_sources(user: dict = Depends(get_current_user)):
    """Source options for the ticket/review upload dropdown (registry-driven)."""
    from ml.connectors.registry import SOURCES
    return {"sources": SOURCES}


@router.get("/uploads/data-status")
def upload_data_status(
    clientId: str = Query(...),
    user: dict = Depends(get_current_user),
):
    """Which master types already have COMMITTED rows for this client.

    Drives the incremental-upload gate: a REQUIRED master counts as satisfied if
    it's already committed (real table has rows), so an existing client can
    upload just one master (e.g. new support tickets) without re-uploading the
    core masters. A brand-new client still has everything False → must upload
    all required masters. Table names come from the trusted internal
    MASTER_TYPE_TO_TABLE mapping (never user input)."""
    _require_client_access(user, clientId)
    committed: Dict[str, bool] = {}
    with engine.connect() as conn:
        for master_type, (table, _cols) in MASTER_TYPE_TO_TABLE.items():
            exists = conn.execute(
                text(f"SELECT EXISTS (SELECT 1 FROM {table} WHERE client_id = :cid)"),
                {"cid": clientId},
            ).scalar()
            committed[master_type] = bool(exists)
    return {"committed": committed}


@router.get("/uploads")
def list_uploads(
    clientId: str = Query(...),
    user: dict = Depends(get_current_user),  # Audit fix #1
):
    """
    List staged files in the client's PENDING batch.

    Queries the database (not an in-memory dict), so results are correct
    even across app restarts. For each master_type that has rows in its
    staging table for the pending batch, returns a row with the count.

    If there is no pending batch, returns an empty list.
    """
    _require_client_access(user, clientId)
    with engine.begin() as conn:
        batch_id = _get_pending_batch_id(conn, clientId)
        if not batch_id:
            return []

        uploads = []
        for master_type, (real_table, _) in MASTER_TYPE_TO_TABLE.items():
            staging_table = f"staging_{real_table}"
            result = conn.execute(
                text(f"SELECT COUNT(*) FROM {staging_table} "
                     f"WHERE batch_id = :bid AND client_id = :cid"),
                {"bid": batch_id, "cid": clientId},
            ).fetchone()
            row_count = result[0] if result else 0
            if row_count > 0:
                # The original file name + upload time live in the file_upload audit
                # event ("<masterType> · <filename> · <N> rows") — the only place the
                # uploaded filename is persisted. Surface them so the UI can show WHICH
                # file was uploaded even after a reload / new session.
                arow = conn.execute(
                    text("SELECT split_part(details, ' · ', 2) AS fname, ts "
                         "FROM audit_log WHERE client_id = :cid "
                         "AND action_type = 'file_upload' "
                         "AND split_part(details, ' · ', 1) = :mt "
                         "ORDER BY ts DESC LIMIT 1"),
                    {"cid": clientId, "mt": master_type},
                ).fetchone()
                file_name = arow[0] if arow and arow[0] else f"{master_type}.csv"
                uploaded_at = (arow[1].isoformat() if arow and arow[1]
                               else datetime.now().isoformat())
                uploads.append({
                    "masterType": master_type,
                    "fileName": file_name,
                    "fileSize": 0,            # original byte size isn't persisted
                    "rowCount": row_count,
                    "uploadedAt": uploaded_at,
                    # 'success', NOT 'staged': the frontend UploadStatus contract is
                    # idle|uploading|success|error. A staged file IS successfully
                    # uploaded (pending commit); returning 'staged' made the per-card
                    # state, the "N uploaded" counter, and the Save button all treat
                    # the file as absent on reload.
                    "status": "success",
                    "stagingTable": staging_table,
                    "batchId": batch_id,
                })
        return uploads


# Safety cap on preview size: enough to show every realistic upload in full
# (the largest master, line_items, is ~25k rows) while preventing a pathological
# file from sending a huge payload / freezing the browser's DOM.
MAX_PREVIEW_ROWS = 50000


@router.get("/uploads/preview")
def preview_upload(
    clientId: str = Query(...),
    masterType: str = Query(...),
    limit: int = Query(MAX_PREVIEW_ROWS, ge=1, le=MAX_PREVIEW_ROWS),
    user: dict = Depends(get_current_user),
):
    """Preview a STAGED master file in the client's pending batch, so the user
    can see WHAT they uploaded before saving.

    By default returns the WHOLE file (up to MAX_PREVIEW_ROWS as a safety cap so
    a pathologically large upload can't freeze the browser). Reads from the
    staging table (the source of truth for unsaved data), returns the user-facing
    columns (internal client_id/batch_id/row-id are hidden) plus the total staged
    row count so the UI can show "all N rows" vs "first N of M".
    """
    _require_client_access(user, clientId)
    if masterType not in MASTER_TYPE_TO_TABLE:
        raise HTTPException(status_code=404, detail=f"Unknown master type: {masterType}")

    real_table, expected_cols = MASTER_TYPE_TO_TABLE[masterType]
    staging_table = f"staging_{real_table}"
    cols = [c for c in expected_cols if c != "client_id"]  # what the user uploaded
    col_list = ", ".join(f'"{c}"' for c in cols)

    with engine.begin() as conn:
        batch_id = _get_pending_batch_id(conn, clientId)
        if not batch_id:
            raise HTTPException(status_code=404, detail="No pending upload to preview.")

        total = conn.execute(
            text(f"SELECT COUNT(*) FROM {staging_table} "
                 f"WHERE batch_id = :bid AND client_id = :cid"),
            {"bid": batch_id, "cid": clientId},
        ).scalar() or 0
        if total == 0:
            raise HTTPException(
                status_code=404,
                detail=f"No staged '{masterType}' file to preview.",
            )

        # ctid ≈ physical insertion order for a freshly-staged batch, so the
        # preview shows the file's first rows in upload order.
        rows = conn.execute(
            text(f"SELECT {col_list} FROM {staging_table} "
                 f"WHERE batch_id = :bid AND client_id = :cid ORDER BY ctid LIMIT :lim"),
            {"bid": batch_id, "cid": clientId, "lim": limit},
        ).fetchall()

        # Original filename lives in the file_upload audit event (same source as
        # list_uploads), so the preview header can name the actual file.
        arow = conn.execute(
            text("SELECT split_part(details, ' · ', 2) FROM audit_log "
                 "WHERE client_id = :cid AND action_type = 'file_upload' "
                 "AND split_part(details, ' · ', 1) = :mt ORDER BY ts DESC LIMIT 1"),
            {"cid": clientId, "mt": masterType},
        ).fetchone()

    file_name = arow[0] if arow and arow[0] else f"{masterType}.csv"
    return {
        "masterType": masterType,
        "fileName": file_name,
        "columns": cols,
        "rows": [list(r) for r in rows],   # FastAPI serializes dates/Decimals
        "shownRows": len(rows),
        "totalRows": total,
    }


@router.get("/uploads/saved-preview")
def saved_preview(
    clientId: str = Query(...),
    masterType: str = Query(...),
    limit: int = Query(MAX_PREVIEW_ROWS, ge=1, le=MAX_PREVIEW_ROWS),
    user: dict = Depends(get_current_user),
):
    """Preview SAVED (committed) master data, so after committing an upload the
    "your data has been saved" banner can show the client WHAT was saved.

    Unlike preview_upload (which reads the pending batch's staging table), this
    reads the REAL table scoped to client_id — there is no batch after a commit.
    Returns the same UploadPreview shape so the frontend renders it with the same
    preview modal. A master with no committed rows returns columns + 0 rows (a
    valid state), NOT a 404 — "no saved rows" is not an error here.
    """
    _require_client_access(user, clientId)
    if masterType not in MASTER_TYPE_TO_TABLE:
        raise HTTPException(status_code=404, detail=f"Unknown master type: {masterType}")

    real_table, expected_cols = MASTER_TYPE_TO_TABLE[masterType]
    cols = [c for c in expected_cols if c != "client_id"]   # hide internal client_id
    col_list = ", ".join(f'"{c}"' for c in cols)

    with engine.connect() as conn:
        total = conn.execute(
            text(f"SELECT COUNT(*) FROM {real_table} WHERE client_id = :cid"),
            {"cid": clientId},
        ).scalar() or 0

        # ctid ≈ physical order — cheap, stable, and good enough for a preview
        # of saved rows (mirrors the staged preview, which orders by ctid too).
        rows = conn.execute(
            text(f"SELECT {col_list} FROM {real_table} "
                 f"WHERE client_id = :cid ORDER BY ctid LIMIT :lim"),
            {"cid": clientId, "lim": limit},
        ).fetchall()

        # Name the header after the file the client last uploaded for this master
        # (same source as preview_upload), falling back to a generic label.
        arow = conn.execute(
            text("SELECT split_part(details, ' · ', 2) FROM audit_log "
                 "WHERE client_id = :cid AND action_type = 'file_upload' "
                 "AND split_part(details, ' · ', 1) = :mt ORDER BY ts DESC LIMIT 1"),
            {"cid": clientId, "mt": masterType},
        ).fetchone()

    file_name = arow[0] if arow and arow[0] else f"{masterType} (saved)"
    return {
        "masterType": masterType,
        "fileName": file_name,
        "columns": cols,
        "rows": [list(r) for r in rows],
        "shownRows": len(rows),
        "totalRows": total,
    }


@router.get("/uploads/batch")
def get_pending_batch(
    clientId: str = Query(...),
    user: dict = Depends(get_current_user),  # Audit fix #1
):
    """
    Return the pending batch summary for a client (or null if none).

    Use this on the Upload page to show "You have N files staged. Ready
    to commit?" before the user clicks the Commit button.
    """
    _require_client_access(user, clientId)
    with engine.begin() as conn:
        row = conn.execute(
            text("SELECT batch_id::text, created_at, status "
                 "FROM upload_batches "
                 "WHERE client_id = :cid AND status = 'pending'"),
            {"cid": clientId},
        ).fetchone()
        if not row:
            return {"pendingBatch": None}

        batch_id, created_at, status = row

        # Count rows per staging table
        per_type = []
        total_rows = 0
        for master_type, (real_table, _) in MASTER_TYPE_TO_TABLE.items():
            staging_table = f"staging_{real_table}"
            count_row = conn.execute(
                text(f"SELECT COUNT(*) FROM {staging_table} "
                     f"WHERE batch_id = :bid AND client_id = :cid"),
                {"bid": batch_id, "cid": clientId},
            ).fetchone()
            n = count_row[0] if count_row else 0
            if n > 0:
                per_type.append({"masterType": master_type, "rowCount": n})
                total_rows += n

        return {
            "pendingBatch": {
                "batchId": batch_id,
                "createdAt": created_at.isoformat() if created_at else None,
                "status": status,
                "totalRows": total_rows,
                "files": per_type,
            }
        }


@router.delete("/uploads/{master_type}")
def delete_upload(
    master_type: str,
    request: Request,
    clientId: str = Query(...),
    user: dict = Depends(get_current_user),  # Audit fix #1
):
    """
    Remove a single master_type's rows from the pending batch.

    Useful if the user staged the wrong file and wants to redo just that
    one. If no pending batch exists, this is a no-op.
    """
    _require_client_access(user, clientId)
    if master_type not in VALID_MASTER_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown master type: {master_type}",
        )

    real_table, _ = MASTER_TYPE_TO_TABLE[master_type]
    staging_table = f"staging_{real_table}"

    with engine.begin() as conn:
        batch_id = _get_pending_batch_id(conn, clientId)
        if not batch_id:
            return {"deleted": 0, "reason": "No pending batch for this client"}

        result = conn.execute(
            text(f"DELETE FROM {staging_table} "
                 f"WHERE batch_id = :bid AND client_id = :cid"),
            {"bid": batch_id, "cid": clientId},
        )
        deleted = result.rowcount

        # If removing this master_type left the batch with zero staging rows
        # across all tables, delete the upload_batches row too. Otherwise
        # /uploads/batch would keep returning a ghost "0-file" pending batch.
        batch_cleaned = _cleanup_empty_batch(conn, batch_id, clientId)

    # Also remove the saved file from disk (best-effort)
    try:
        for f in os.listdir(UPLOAD_DIR):
            if f.startswith(f"{clientId}_{master_type}_"):
                os.remove(os.path.join(UPLOAD_DIR, f))
    except Exception:
        pass

    # Audit fix #4: log the staging-row delete so we can trace
    # "user X removed Y master from pending batch Z" later.
    log_audit_event(
        request,
        action_type="batch_file_remove",
        details=f"batch={batch_id} · {master_type} · {deleted:,} rows removed",
        client_id=clientId,
        user_id=user.get("id"),
        user_email=user.get("email"),
        outcome="success",
    )

    return {
        "deleted": deleted,
        "masterType": master_type,
        "batchId": batch_id,
        "batchCleaned": batch_cleaned,
    }


# ── Commit / discard endpoints ───────────────────────────────────────────────

def _preflight_fk_check(conn, batch_id: str, client_id: str) -> list[dict]:
    """
    Check every FK rule in CATALOG_FK_CHECKS.

    For each rule, a staging row is valid if its foreign key value exists
    EITHER in the real parent table (already committed) OR in the parent's
    own staging table for the SAME batch (being committed together).

    Returns a list of violation dicts. Empty list = all clean.
    """
    violations = []

    for (child_staging, child_col, parent_real, parent_real_col,
         parent_staging, fk_name) in CATALOG_FK_CHECKS:

        # Find staging rows whose FK value is missing from both the real
        # parent table AND the parent's staging rows for this batch.
        query = text(f"""
            SELECT s.{child_col} AS missing_key, COUNT(*) AS row_count
            FROM {child_staging} s
            WHERE s.batch_id = :bid
              AND s.client_id = :cid
              AND s.{child_col} IS NOT NULL
              AND NOT EXISTS (
                  SELECT 1 FROM {parent_real} p
                  WHERE p.client_id = :cid
                    AND p.{parent_real_col} = s.{child_col}
              )
              AND NOT EXISTS (
                  SELECT 1 FROM {parent_staging} ps
                  WHERE ps.batch_id = :bid
                    AND ps.client_id = :cid
                    AND ps.{parent_real_col} = s.{child_col}
              )
            GROUP BY s.{child_col}
            LIMIT 10
        """)
        rows = conn.execute(query, {"bid": batch_id, "cid": client_id}).fetchall()
        for missing_key, row_count in rows:
            violations.append({
                "fk": fk_name,
                "stagingTable": child_staging,
                "column": child_col,
                "missingKey": missing_key,
                "rowCount": row_count,
                "parentTable": parent_real,
            })

    return violations


# Real/staging table name → the file label the client sees in the upload UI.
_TABLE_FRIENDLY = {
    "customers": "Customer", "orders": "Order", "line_items": "Line Items",
    "products": "Product", "product_prices": "Product Price",
    "product_vendor_mapping": "Product-Vendor Mapping",
    "categories": "Category", "sub_categories": "Sub-Category",
    "sub_sub_categories": "Sub-Sub-Category", "brands": "Brand",
    "vendors": "Vendor", "customer_reviews": "Customer Reviews",
    "support_tickets": "Support Tickets",
}


def _humanize_fk_violations(violations: list[dict]) -> str:
    """Turn raw preflight violations into a plain-English, actionable message.

    The raw violation dicts (internal fk/staging-table names, 'missingKey') are
    useless to a client. This names the client's own files and the offending
    value, and says exactly how to fix it. The raw list is logged server-side
    for diagnostics.
    """
    def friendly(table: str) -> str:
        base = table[len("staging_"):] if table.startswith("staging_") else table
        return _TABLE_FRIENDLY.get(base, base)

    lines = []
    for v in violations:
        child = friendly(v.get("stagingTable", ""))
        parent = friendly(v.get("parentTable", ""))
        key = v.get("missingKey")
        rc = v.get("rowCount", 0)
        rows = "row" if rc == 1 else "rows"
        lines.append(
            f"Your {child} file references {parent} \"{key}\", but no such "
            f"{parent} exists in your {parent} file or your saved data ({rc} {rows}). "
            f"Add {parent} {key} to your {parent} file, or remove the {child} "
            f"rows that reference it."
        )

    if not lines:
        return "Some files reference items that aren't in your upload or saved data."
    if len(lines) == 1:
        return lines[0]
    return ("Some files reference items that aren't in your upload or saved data:\n"
            "• " + "\n• ".join(lines))


def _commit_batch(conn, batch_id: str, client_id: str) -> dict:
    """
    Move rows from staging → real tables in COMMIT_ORDER.

    The caller must have already opened a transaction and run
    SET CONSTRAINTS ALL DEFERRED — so FK checks on transactional tables
    (orders, line_items, etc.) happen at COMMIT, not after each INSERT.

    Each INSERT uses ON CONFLICT handling driven by COMMIT_CONFLICT:
      - dimension tables get UPSERT so re-uploads refresh non-PK columns
      - fact tables get DO NOTHING so historical records aren't overwritten
    This makes commit replay-safe: a batch that partially conflicts with
    existing data no longer blows up the whole transaction.
    """
    per_type_rows = {}
    mv_needs_refresh = False

    for master_type in COMMIT_ORDER:
        real_table, expected_cols = MASTER_TYPE_TO_TABLE[master_type]
        staging_table = f"staging_{real_table}"
        pk_cols, action = COMMIT_CONFLICT[master_type]

        # Only insert columns that exist in the real table (skip batch_id, staging_row_id)
        col_list = ", ".join(expected_cols)
        pk_list = ", ".join(pk_cols)

        # Build ON CONFLICT clause based on strategy for this master_type.
        if action == "update":
            # Update every non-PK column to the staging value (EXCLUDED.col).
            non_pk_cols = [c for c in expected_cols if c not in pk_cols]
            if non_pk_cols:
                set_clause = ", ".join(
                    f"{c} = EXCLUDED.{c}" for c in non_pk_cols
                )
                conflict_sql = f"ON CONFLICT ({pk_list}) DO UPDATE SET {set_clause}"
            else:
                # Table has no non-PK columns — nothing to update, just skip.
                conflict_sql = f"ON CONFLICT ({pk_list}) DO NOTHING"
        else:  # 'nothing'
            conflict_sql = f"ON CONFLICT ({pk_list}) DO NOTHING"

        insert_sql = text(f"""
            INSERT INTO {real_table} ({col_list})
            SELECT {col_list}
            FROM {staging_table}
            WHERE batch_id = :bid AND client_id = :cid
            {conflict_sql}
        """)
        result = conn.execute(insert_sql, {"bid": batch_id, "cid": client_id})
        # rowcount counts affected rows (inserts + updates for UPSERT; inserts
        # only for DO NOTHING). Use it as the "rows committed" metric.
        affected = result.rowcount
        per_type_rows[master_type] = affected

        if affected > 0 and real_table in MV_TRIGGER_TABLES:
            mv_needs_refresh = True

    # Clear this batch's staging rows (other pending batches — none for this
    # client — are unaffected because we filter by batch_id)
    for master_type in COMMIT_ORDER:
        real_table, _ = MASTER_TYPE_TO_TABLE[master_type]
        staging_table = f"staging_{real_table}"
        conn.execute(
            text(f"DELETE FROM {staging_table} WHERE batch_id = :bid"),
            {"bid": batch_id},
        )

    # Mark the batch committed
    conn.execute(
        text("UPDATE upload_batches SET status = 'committed', "
             "committed_at = NOW() WHERE batch_id = :bid"),
        {"bid": batch_id},
    )

    return {"perTypeRows": per_type_rows, "mvNeedsRefresh": mv_needs_refresh}


# NOTE: The @router.post("/uploads/commit") and @router.post("/uploads/discard")
# endpoint decorators used to live here. They were moved UP to appear before
# @router.post("/uploads/{master_type}") because FastAPI matches routes in
# declaration order — the parameterized path was catching /uploads/commit
# requests and treating "commit" as a master_type. The helper functions
# _preflight_fk_check() and _commit_batch() remain above; Python resolves
# them at call time, so their position doesn't matter.
