"""
fix_cascade_reload.py — Reload tables wiped by CASCADE truncate
================================================================
When fix_categories.py ran TRUNCATE categories CASCADE, it also wiped:
  - sub_categories (FK → categories)
  - sub_sub_categories (FK → categories)
  - products (FK → categories, sub_categories, sub_sub_categories)
  - product_prices (FK → products)
  - product_vendor_mapping (FK → products)
  - line_items (FK → products)

Categories/sub_categories/sub_sub_categories were already reloaded.
This script reloads: products, product_prices, product_vendor_mapping,
and line_items from the raw Excel file.

Usage:
    cd ~/Desktop/analyst_agent_v3
    source venv/bin/activate
    python db/fix_cascade_reload.py \
        --excel ~/Desktop/analyst_agent_v3/db/walmart_raw_data_template_v6.xlsx \
        --db-url postgresql://chaitanya:@localhost:5432/walmart_crp
"""

import argparse
import logging
import openpyxl
import psycopg2
from datetime import datetime, date

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("fix_cascade")


def clean(val):
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip()
        if v.startswith("="):
            if "TRUE" in v.upper(): return True
            if "FALSE" in v.upper(): return False
            return None
        if v.startswith("──") or v == "":
            return None
        return v
    if isinstance(val, datetime):
        return val.date() if val.hour == val.minute == val.second == 0 else val
    return val


def find_header_row(rows, expected_col):
    expected_lower = expected_col.lower().strip()
    for i, row in enumerate(rows):
        for cell in row:
            if cell is not None and str(cell).strip().lower() == expected_lower:
                return i
    raise ValueError(f"Header row with column '{expected_col}' not found")


def get_data_rows(rows, header_idx, num_cols):
    result = []
    for row in rows[header_idx + 1:]:
        cleaned = tuple(clean(v) for v in row[:num_cols])
        if cleaned[0] is None:
            continue
        result.append(cleaned)
    return result


def match_sheet_name(wb_sheets, key):
    key_lower = key.lower().strip()
    exact, endswith, contains = [], [], []
    for name in wb_sheets:
        name_lower = name.strip().lower()
        if name_lower == key_lower:
            exact.append(name)
        elif name_lower.endswith(key_lower):
            endswith.append(name)
        elif key_lower in name_lower:
            contains.append(name)
    for bucket in (exact, endswith, contains):
        if bucket:
            return min(bucket, key=len)
    return None


def load_and_insert(cur, wb, sheet_key, table_name, header_col, columns, num_cols):
    """Generic load from Excel sheet and insert into table."""
    actual_sheet = match_sheet_name(wb.sheetnames, sheet_key)
    if actual_sheet is None:
        log.warning("Sheet not found: '%s' — skipping", sheet_key)
        return 0

    log.info("Loading: %s → %s", actual_sheet, table_name)
    ws = wb[actual_sheet]
    rows = list(ws.iter_rows(values_only=True))
    hi = find_header_row(rows, header_col)
    data = get_data_rows(rows, hi, num_cols)

    placeholders = ", ".join(["%s"] * num_cols)
    col_names = ", ".join(columns)
    sql = f"INSERT INTO {table_name} ({col_names}) VALUES ({placeholders}) ON CONFLICT DO NOTHING"

    for row in data:
        # Convert booleans to int for smallint columns
        converted = []
        for v in row:
            if isinstance(v, bool):
                converted.append(1 if v else 0)
            else:
                converted.append(v)
        cur.execute(sql, tuple(converted))

    log.info("  Inserted %d rows into %s", len(data), table_name)
    return len(data)


def main():
    parser = argparse.ArgumentParser(description="Reload cascade-wiped tables from raw Excel")
    parser.add_argument("--excel", required=True, help="Path to raw Excel file")
    parser.add_argument("--db-url", required=True, help="PostgreSQL connection URL")
    args = parser.parse_args()

    log.info("Loading workbook: %s", args.excel)
    wb = openpyxl.load_workbook(args.excel, read_only=True)

    conn = psycopg2.connect(args.db_url)
    cur = conn.cursor()

    # Check current state
    tables_to_check = ["categories", "sub_categories", "sub_sub_categories",
                       "products", "product_prices", "product_vendor_mapping", "line_items"]
    log.info("\n=== BEFORE state ===")
    for tbl in tables_to_check:
        cur.execute(f"SELECT count(*) FROM {tbl}")
        cnt = cur.fetchone()[0]
        log.info("  %s: %d rows", tbl, cnt)

    # Disable FK checks
    cur.execute("SET session_replication_role = replica;")

    # 1. Products
    load_and_insert(cur, wb, "Product Master", "products", "product_id",
                    ["product_id", "sku", "product_name", "category_id", "sub_category_id",
                     "sub_sub_category_id", "brand_id", "product_price_id", "rating", "active", "not_available"],
                    11)
    conn.commit()

    # 2. Product Prices
    load_and_insert(cur, wb, "Product Price Master", "product_prices", "price_id",
                    ["price_id", "product_id", "qty_range_label", "qty_min", "qty_max", "unit_price_usd"],
                    6)
    conn.commit()

    # 3. Product-Vendor Mapping
    load_and_insert(cur, wb, "Product-Vendor Mapping", "product_vendor_mapping", "pv_id",
                    ["pv_id", "product_id", "brand_id", "vendor_id"],
                    4)
    conn.commit()

    # 4. Line Items
    load_and_insert(cur, wb, "Line Items Master", "line_items", "line_item_id",
                    ["client_id", "line_item_id", "order_id", "customer_id", "product_id",
                     "quantity", "unit_price_usd", "final_line_total_usd", "item_discount_usd", "item_status"],
                    10)
    conn.commit()

    # Re-enable FK checks
    cur.execute("SET session_replication_role = DEFAULT;")
    conn.commit()

    # Show AFTER state
    log.info("\n=== AFTER state ===")
    for tbl in tables_to_check:
        cur.execute(f"SELECT count(*) FROM {tbl}")
        cnt = cur.fetchone()[0]
        log.info("  %s: %d rows", tbl, cnt)

    # Refresh materialized view
    try:
        log.info("\nRefreshing materialized view...")
        cur.execute("REFRESH MATERIALIZED VIEW mv_customer_features;")
        conn.commit()
        log.info("✅ Materialized view refreshed")
    except Exception as e:
        conn.rollback()
        log.warning("Materialized view refresh failed: %s", e)

    cur.close()
    conn.close()
    log.info("\n✅ All cascade-affected tables reloaded!")


if __name__ == "__main__":
    main()
