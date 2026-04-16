"""
fix_categories.py — Fix category/sub-category/sub-sub-category data collision
==============================================================================
The match_sheet_name() bug in load_data.py caused all 3 category tables
to be loaded with sub_sub_category data. This script re-imports the correct
data from the raw Excel file.

Usage:
    cd ~/Desktop/analyst_agent_v3
    source venv/bin/activate
    python db/fix_categories.py --excel <path_to_excel> --db-url postgresql://chaitanya:@localhost:5432/walmart_crp
"""

import argparse
import logging
import openpyxl
import psycopg2

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("fix_categories")


def clean(val):
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip()
        if v.startswith("=") or v.startswith("──") or v == "":
            return None
        return v
    return val


def find_header_row(rows, expected_col):
    expected_lower = expected_col.lower().strip()
    for i, row in enumerate(rows):
        for cell in row:
            if cell is not None and str(cell).strip().lower() == expected_lower:
                return i
    raise ValueError(f"Header row with column '{expected_col}' not found")


def get_data_rows(rows, header_idx, num_cols):
    result = []
    for row in rows[header_idx + 1:]:
        cleaned = tuple(clean(v) for v in row[:num_cols])
        if cleaned[0] is None:
            continue
        result.append(cleaned)
    return result


def match_sheet_name(wb_sheets, key):
    """Fixed version — picks shortest match to avoid 'Category Master' hitting 'Sub-Sub-Category Master'."""
    key_lower = key.lower().strip()
    exact, endswith, contains = [], [], []
    for name in wb_sheets:
        name_lower = name.strip().lower()
        if name_lower == key_lower:
            exact.append(name)
        elif name_lower.endswith(key_lower):
            endswith.append(name)
        elif key_lower in name_lower:
            contains.append(name)
    for bucket in (exact, endswith, contains):
        if bucket:
            return min(bucket, key=len)
    return None


def main():
    parser = argparse.ArgumentParser(description="Fix category tables from raw Excel")
    parser.add_argument("--excel", required=True, help="Path to raw Excel file")
    parser.add_argument("--db-url", required=True, help="PostgreSQL connection URL")
    args = parser.parse_args()

    log.info("Loading workbook: %s", args.excel)
    wb = openpyxl.load_workbook(args.excel, read_only=True)

    # Match correct sheets
    cat_sheet = match_sheet_name(wb.sheetnames, "Category Master")
    sub_sheet = match_sheet_name(wb.sheetnames, "Sub-Category Master")
    subsub_sheet = match_sheet_name(wb.sheetnames, "Sub-Sub-Category Master")

    log.info("Category sheet:         %s", cat_sheet)
    log.info("Sub-Category sheet:     %s", sub_sheet)
    log.info("Sub-Sub-Category sheet: %s", subsub_sheet)

    # ── Parse categories ──
    ws1 = wb[cat_sheet]
    rows1 = list(ws1.iter_rows(values_only=True))
    hi1 = find_header_row(rows1, "category_id")
    cat_data = get_data_rows(rows1, hi1, 2)
    log.info("Categories parsed: %d rows", len(cat_data))
    for r in cat_data:
        log.info("  %s → %s", r[0], r[1])

    # ── Parse sub_categories ──
    ws2 = wb[sub_sheet]
    rows2 = list(ws2.iter_rows(values_only=True))
    hi2 = find_header_row(rows2, "sub_category_id")
    sub_data = get_data_rows(rows2, hi2, 3)
    log.info("Sub-categories parsed: %d rows", len(sub_data))

    # ── Parse sub_sub_categories ──
    ws3 = wb[subsub_sheet]
    rows3 = list(ws3.iter_rows(values_only=True))
    hi3 = find_header_row(rows3, "sub_sub_category_id")
    subsub_data = get_data_rows(rows3, hi3, 4)
    log.info("Sub-sub-categories parsed: %d rows", len(subsub_data))

    # ── Write to database ──
    conn = psycopg2.connect(args.db_url)
    cur = conn.cursor()

    # Show BEFORE state
    for tbl in ("categories", "sub_categories", "sub_sub_categories"):
        cur.execute(f"SELECT count(*) FROM {tbl}")
        cnt = cur.fetchone()[0]
        cur.execute(f"SELECT * FROM {tbl} LIMIT 3")
        sample = cur.fetchall()
        log.info("BEFORE %s: %d rows, sample: %s", tbl, cnt, sample)

    # Disable FK checks
    cur.execute("SET session_replication_role = replica;")

    # Truncate and reload
    cur.execute("TRUNCATE TABLE categories CASCADE;")
    for row in cat_data:
        cur.execute("INSERT INTO categories (category_id, category_name) VALUES (%s, %s)", row)
    log.info("✅ Inserted %d categories", len(cat_data))

    cur.execute("TRUNCATE TABLE sub_categories CASCADE;")
    for row in sub_data:
        cur.execute("INSERT INTO sub_categories (sub_category_id, sub_category_name, category_id) VALUES (%s, %s, %s)", row)
    log.info("✅ Inserted %d sub_categories", len(sub_data))

    cur.execute("TRUNCATE TABLE sub_sub_categories CASCADE;")
    for row in subsub_data:
        cur.execute("INSERT INTO sub_sub_categories (sub_sub_category_id, sub_sub_category_name, sub_category_id, category_id) VALUES (%s, %s, %s, %s)", row)
    log.info("✅ Inserted %d sub_sub_categories", len(subsub_data))

    # Re-enable FK checks
    cur.execute("SET session_replication_role = DEFAULT;")

    conn.commit()

    # Show AFTER state
    for tbl in ("categories", "sub_categories", "sub_sub_categories"):
        cur.execute(f"SELECT count(*) FROM {tbl}")
        cnt = cur.fetchone()[0]
        cur.execute(f"SELECT * FROM {tbl} LIMIT 5")
        sample = cur.fetchall()
        log.info("AFTER %s: %d rows, sample: %s", tbl, cnt, sample)

    # Refresh materialized view
    try:
        cur.execute("REFRESH MATERIALIZED VIEW mv_customer_features;")
        conn.commit()
        log.info("✅ Materialized view refreshed")
    except Exception as e:
        conn.rollback()
        log.warning("Materialized view refresh failed: %s", e)

    cur.close()
    conn.close()
    log.info("✅ Done! All three category tables fixed.")


if __name__ == "__main__":
    main()
