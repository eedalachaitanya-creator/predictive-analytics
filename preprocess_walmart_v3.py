"""
preprocess_walmart_v3.py  —  Stage 2 of the Walmart Agentic AI Pipeline
========================================================================
Input  : walmart_raw_data_template_v5.xlsx  (15-sheet normalised template)
Output : <prefix>_clean_ai_dataset_v7.xlsx  (derived feature-engineering output)

Pipeline Stages
  Stage 1  — INGEST          Read all 15 sheets from v5 template
  Stage 2  — CONFIG          Load client parameters + wire all master sheets
  Stage 3  — VALIDATE        Null checks, duplicate detection, type enforcement
  Stage 4  — NORMALISE       Standardise dates, amounts, text formats
  Stage 5  — IDENTITY        (client_id, customer_id) composite key + email index
  Stage 6  — REPEAT          Order frequency, gaps, retention labels
  Stage 7  — HIGH VALUE      Tier assignment driven by Value-Tier Master (dynamic)
  Stage 8  — PRODUCT AGGS    Category / brand / vendor aggregations + segment rules
  Stage 9  — FEATURE ENG     Build ML feature vector + Value Proposition lookup
  Stage 10 — OUTPUT          Write clean_ai_dataset_v7.xlsx (12 sheets)

v3 Schema Changes (v4 → v5 upgrade)
  ✅ client_id   replaces vendor_id  as the retail-client identifier in all tx tables
                 vendor_id is now reserved for product-supply vendors (Vendor Master)
  ✅ Vendor Master added  (45 suppliers — companies that supply products to clients)
  ✅ Product-Vendor Mapping added  (product ↔ vendor ↔ brand bridge table)
  ✅ Product Price Master  normalised row-wise  (5 qty tiers per product)
               pipeline extracts the 1-unit tier as the base price
  ✅ Brand Master  expanded to 156 brands with active / not_available / category_hint
  ✅ Product Master  active / not_available columns added
  ✅ Config parameters renamed: vendor_name → client_name, vendor_code → client_code,
                                vendor_id   → client_id
  ✅ New ML features: vendor_diversity, brand_active_pct, active_brand_purchases
  ✅ New output sheet: 🏭 Vendor Analysis  (per-vendor sales contribution)

v2 Baseline Features (retained unchanged)
  ✅ Median/recent-gap personal_rhythm_breached  (not mean — resistant to outliers)
  ✅ Value-Tier Master drives tier assignment  (quartile or fixed method)
  ✅ Business Segment Master drives segment assignment  (eval-safe rule engine)
  ✅ Value Proposition Master drives per-customer action plan  (tier × risk lookup)
  ✅ Multi-tenant architecture  —  CLIENT_CODE scopes all outputs
  ✅ Zero hardcoded rules  —  all thresholds / tiers / segments driven by master sheets
"""

import os
import sys
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# ─────────────────────────────────────────────────────────────────────────────
# PATH RESOLUTION (multi-tenant — pass CLIENT_DIR as argv[1] or run in-place)
# ─────────────────────────────────────────────────────────────────────────────
if len(sys.argv) >= 2:
    CLIENT_DIR = sys.argv[1]
else:
    CLIENT_DIR = os.path.dirname(os.path.abspath(__file__))

CLIENT_DIR  = os.path.abspath(CLIENT_DIR)
INPUT_FILE  = os.path.join(CLIENT_DIR, "walmart_raw_data_template_v5.xlsx")
OUTPUT_DIR  = os.path.join(CLIENT_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

if not os.path.exists(INPUT_FILE):
    raise SystemExit(
        f"\n❌  Input file not found: {INPUT_FILE}\n"
        f"    Place walmart_raw_data_template_v5.xlsx inside: {CLIENT_DIR}\n"
    )

OUTPUT_FILE = None  # set in Stage 2 after CLIENT_CODE is known


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def log(stage, msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] STAGE {stage} | {msg}")


audit_log = []

def audit(stage, key, value):
    audit_log.append({"stage": stage, "check": key, "result": str(value),
                      "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(wb, title, df, header_color, tab_color,
                freeze="A2", col_widths=None, banner_subtitle=""):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color.lstrip("#")

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=max(len(df.columns), 1))
    bc = ws.cell(1, 1, f"{title}   |   {banner_subtitle}")
    bc.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    bc.fill      = PatternFill("solid", start_color=tab_color.lstrip("#"))
    bc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    hf = PatternFill("solid", start_color=header_color.lstrip("#"))
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(2, c, col)
        cell.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill      = hf
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[2].height = 26

    alt = PatternFill("solid", start_color="F7F9FC")
    wht = PatternFill("solid", start_color="FFFFFF")
    for r, row in enumerate(df.itertuples(index=False), 3):
        fill = alt if r % 2 == 1 else wht
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c)
            cell.value     = "" if (not isinstance(val, str) and pd.isna(val)) else val
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = Alignment(vertical="center")

    if col_widths:
        for ltr, w in col_widths.items():
            ws.column_dimensions[ltr].width = w
    else:
        for c, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)),
                          max((len(str(v)) for v in df.iloc[:, c - 1]
                               if pd.notna(v)), default=0))
            ws.column_dimensions[get_column_letter(c)].width = min(max_len + 3, 45)

    ws.freeze_panes = freeze
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 1 — INGEST
# ══════════════════════════════════════════════════════════════════════════════
log(1, f"Reading input file: {INPUT_FILE}")


def rsheet(name, skip=1):
    """Read a v5 sheet.  Most sheets: title row 1, headers row 2 → skip=1.
       Product Price Master has an extra note row: title / note / headers → skip=2.
    """
    return pd.read_excel(INPUT_FILE, sheet_name=name, header=0, skiprows=skip)


# — Transaction tables ────────────────────────────────────────────────────────
raw_customers  = rsheet("👤 Customer Master")
raw_orders     = rsheet("📦 Order Master")
raw_lineitems  = rsheet("🛍️ Line Items Master")

# — Config & analytics masters ────────────────────────────────────────────────
raw_config     = rsheet("⚙️ Vendor Config")
raw_segments   = rsheet("💼 Business Segment Master")
raw_tiers      = rsheet("💎 Value-Tier Master")
raw_valueprops = rsheet("🎯 Value Proposition Master")

# — Product catalogue  (3-level hierarchy, Brand, Price, Vendor) ───────────────
raw_categories    = rsheet("📂 Category Master")
raw_subcategories = rsheet("📁 Sub-Category Master")
raw_subsubcats    = rsheet("📄 Sub-Sub-Category Master")
raw_brands        = rsheet("🏷️ Brand Master")   # v5: has active / not_available / category_hint
raw_vendors       = rsheet("🏭 Vendor Master")   # NEW in v5: product supply vendors
raw_pv_mapping    = rsheet("🔗 Product-Vendor Mapping")  # NEW in v5
raw_products      = rsheet("📋 Product Master")  # v5: has active / not_available columns

# Product Price Master: v5 row-wise (5 tiers per product).
# Title row 1, extra note row 2, actual headers row 3 → skip 2 rows.
raw_prices = rsheet("💲 Product Price Master", skip=2)
# Drop any leftover None column
raw_prices = raw_prices.loc[:, raw_prices.columns.notna()]


# — Build base price lookup (1-unit tier = base price) ─────────────────────────
# v5 price tiers: qty_range_label ∈ ['1 unit', '2 – 10', '11 – 50', '51 – 100', '100+']
_base = raw_prices[raw_prices["qty_range_label"] == "1 unit"][
    ["product_id", "unit_price_usd"]
].rename(columns={"unit_price_usd": "base_price_usd"}).drop_duplicates("product_id").copy()

if len(_base) == 0:   # fallback: take the lowest qty_min tier per product
    log(1, "⚠️  No '1 unit' tier found in Product Price Master — using lowest qty_min tier")
    _base = (raw_prices.sort_values("qty_min")
             .groupby("product_id").first()
             .reset_index()[["product_id", "unit_price_usd"]]
             .rename(columns={"unit_price_usd": "base_price_usd"}))


# — Dimension joins → product_enriched ─────────────────────────────────────────
_cat      = raw_categories[["category_id", "category_name"]].copy()
_subcat   = raw_subcategories[["sub_category_id", "sub_category_name"]].rename(
                columns={"sub_category_name": "sub_category"}).copy()
_subsubcat = raw_subsubcats[["sub_sub_category_id", "sub_sub_category_name"]].rename(
                columns={"sub_sub_category_name": "sub_sub_category"}).copy()

# v5 Brand Master: brand_id, brand_name, brand_description, vendor_id, active, not_available, category_hint
_brand = raw_brands[["brand_id", "brand_name", "active", "category_hint"]].rename(
    columns={"brand_name": "brand", "active": "brand_active"}).copy()

# v5 Product-Vendor Mapping: one row per product (use first vendor per product for enrichment)
_pv = raw_pv_mapping[["product_id", "vendor_id"]].drop_duplicates("product_id").copy()
_vendor_name = raw_vendors[["vendor_id", "vendor_name"]].rename(
    columns={"vendor_name": "product_vendor"}).copy()

product_enriched = (
    raw_products
    .merge(_cat,        on="category_id",        how="left")
    .merge(_subcat,     on="sub_category_id",     how="left")
    .merge(_subsubcat,  on="sub_sub_category_id", how="left")
    .merge(_brand,      on="brand_id",            how="left")
    .merge(_base,       on="product_id",          how="left")
    .merge(_pv,         on="product_id",          how="left")
    .merge(_vendor_name,on="vendor_id",           how="left")
)
# product_enriched now includes: product_id, sku, product_name, category_id,
#   sub_category_id, sub_sub_category_id, brand_id, product_price_id,
#   rating, active (product), not_available,
#   category_name, sub_category, sub_sub_category,
#   brand, brand_active, category_hint, base_price_usd, vendor_id, product_vendor

audit(1, "customers_loaded",     len(raw_customers))
audit(1, "orders_loaded",        len(raw_orders))
audit(1, "lineitems_loaded",     len(raw_lineitems))
audit(1, "products_loaded",      len(raw_products))
audit(1, "categories_loaded",    len(raw_categories))
audit(1, "subcategories_loaded", len(raw_subcategories))
audit(1, "subsubcats_loaded",    len(raw_subsubcats))
audit(1, "brands_loaded",        len(raw_brands))
audit(1, "vendors_loaded",       len(raw_vendors))
audit(1, "pv_mappings_loaded",   len(raw_pv_mapping))
audit(1, "price_rows_loaded",    len(raw_prices))
audit(1, "base_prices_extracted",len(_base))

log(1, f"Loaded — customers:{len(raw_customers)}  orders:{len(raw_orders)}  "
       f"lineitems:{len(raw_lineitems)}  products:{len(raw_products)}  "
       f"brands:{len(raw_brands)} ({int(raw_brands['active'].eq(1).sum())} active)  "
       f"vendors:{len(raw_vendors)}  pv_mappings:{len(raw_pv_mapping)}  "
       f"price_tiers:{len(raw_prices)}  base_prices:{len(_base)}  "
       f"cat-hierarchy: {len(raw_categories)} / {len(raw_subcategories)} / {len(raw_subsubcats)}")


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 2 — CONFIG
#
# v5 naming: client_name / client_code / client_id
#   (with fallback to vendor_name / vendor_code / vendor_id for backward compat)
# ══════════════════════════════════════════════════════════════════════════════
log(2, "Loading client configuration and wiring master sheets")

cfg_clean = raw_config[raw_config.iloc[:, 0].notna()].copy()
cfg_clean = cfg_clean[~cfg_clean.iloc[:, 0].astype(str).str.startswith("──")]
# Normalise: handle both "Parameter/Value" (v5) and "parameter/value" column names
cfg_clean.columns = [str(c).strip().lower() for c in cfg_clean.columns]
cfg = dict(zip(cfg_clean["parameter"].str.strip(), cfg_clean["value"]))

# ── v5 client identity (with v4 fallbacks) ────────────────────────────────────
CLIENT_NAME   = str(cfg.get("client_name", cfg.get("vendor_name", "Client")))
CLIENT_CODE   = str(cfg.get("client_code", cfg.get("vendor_code", "CLT")))
CLIENT_ID     = str(cfg.get("client_id",   cfg.get("vendor_id",   f"CLT-{CLIENT_CODE}")))
VENDOR_NAME   = CLIENT_NAME    # alias used in banner strings throughout Stage 10

CURRENCY      = str(cfg.get("report_currency",             "USD"))
CHURN_DAYS    = int(cfg.get("churn_inactivity_days",        90))
RHYTHM_MULT   = float(cfg.get("personal_rhythm_multiplier", 1.5))
REACT_MULT    = float(cfg.get("reactivation_multiplier",    4.0))
REPEAT_THRESHOLD = int(cfg.get("repeat_order_threshold",    2))
OUTPUT_PREFIX = str(cfg.get("output_file_prefix", CLIENT_CODE))
MIN_ORDERS_ML = int(cfg.get("min_orders_for_ml",            2))

ref_raw = str(cfg.get("reference_date", "TODAY")).strip()
REFERENCE_DATE = (pd.Timestamp("today").normalize()
                  if ref_raw.upper() == "TODAY"
                  else pd.Timestamp(ref_raw))

OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"{OUTPUT_PREFIX}_clean_ai_dataset_v7.xlsx")

# Vendor Config tier params — fallback if master sheet is insufficient
_vc_tier_method = str(cfg.get("tier_method", "quartile")).lower()
_vc_tier_names  = [str(cfg.get(f"tier_{i}_name", d))
                   for i, d in [(1, "Platinum"), (2, "Gold"), (3, "Silver"), (4, "Bronze")]]
_vc_fixed_t1    = float(cfg.get("fixed_tier1_min_spend_usd", 500))
_vc_fixed_t2    = float(cfg.get("fixed_tier2_min_spend_usd", 250))
_vc_fixed_t3    = float(cfg.get("fixed_tier3_min_spend_usd", 100))

log(2, f"Client ID: {CLIENT_ID}  |  Code: {CLIENT_CODE}  |  Name: {CLIENT_NAME}  |  "
       f"Churn: {CHURN_DAYS}d  |  Ref date: {REFERENCE_DATE.date()}")
log(2, f"Output: {OUTPUT_FILE}")
audit(2, "client_id",      CLIENT_ID)
audit(2, "client_code",    CLIENT_CODE)
audit(2, "client_name",    CLIENT_NAME)
audit(2, "churn_days",     CHURN_DAYS)
audit(2, "reference_date", str(REFERENCE_DATE.date()))


# ── FIX 1 — Value-Tier Master ────────────────────────────────────────────────
tier_master = raw_tiers.copy()
tier_master.columns = ["tier_id", "tier_name", "threshold_type",
                        "threshold_value", "description", "benefits"]
tier_master = tier_master[tier_master["tier_id"].notna()].copy()
tier_master = tier_master[~tier_master["tier_id"].astype(str).str.startswith("──")]
tier_master["tier_id"]         = tier_master["tier_id"].astype(str).str.strip()
tier_master["threshold_value"] = pd.to_numeric(tier_master["threshold_value"], errors="coerce").fillna(0)
tier_master["tier_name"]       = tier_master["tier_name"].astype(str).str.strip()
tier_master["threshold_type"]  = tier_master["threshold_type"].astype(str).str.lower().str.strip()
tier_master = tier_master[tier_master["tier_name"].notna() & (tier_master["tier_id"] != "nan")]
tier_master = tier_master.sort_values("tier_id").reset_index(drop=True)

if len(tier_master) >= 2:
    TIER_NAMES      = tier_master["tier_name"].tolist()
    first_type      = tier_master["threshold_type"].iloc[0]
    TIER_METHOD     = first_type if first_type in ["quartile", "fixed"] else _vc_tier_method
    TIER_THRESHOLDS = tier_master["threshold_value"].tolist()
    log(2, f"✅ FIX 1 — Value-Tier Master wired | "
           f"names: {TIER_NAMES} | method: {TIER_METHOD} | thresholds: {TIER_THRESHOLDS}")
    audit(2, "tier_source",     "Value-Tier Master (dynamic)")
    audit(2, "tier_method",     TIER_METHOD)
    audit(2, "tier_names",      str(TIER_NAMES))
    audit(2, "tier_thresholds", str(TIER_THRESHOLDS))
else:
    TIER_NAMES      = _vc_tier_names
    TIER_METHOD     = "fixed"
    TIER_THRESHOLDS = [_vc_fixed_t1, _vc_fixed_t2, _vc_fixed_t3, 0]
    log(2, f"⚠️  Value-Tier Master < 2 rows — falling back to Vendor Config")
    audit(2, "tier_source", "Vendor Config fallback")


# ── FIX 2 — Business Segment Master ──────────────────────────────────────────
seg_master = raw_segments.copy()
seg_master.columns = ["segment_id", "segment_name", "description", "criteria", "recommended_focus"]
seg_master = seg_master[seg_master["segment_name"].notna()].copy()
seg_master = seg_master[~seg_master["segment_name"].astype(str).str.startswith("──")]
seg_master["segment_name"] = seg_master["segment_name"].astype(str).str.strip()
seg_master["criteria"]     = seg_master["criteria"].astype(str).str.strip()
seg_master["segment_id"]   = pd.to_numeric(seg_master["segment_id"], errors="coerce")
seg_master = seg_master.sort_values("segment_id").reset_index(drop=True)

SEG_RULES = [(row["segment_name"], row["criteria"]) for _, row in seg_master.iterrows()]
_EMPTY_CRITERIA = {"", "nan", "none", "default", "n/a", "na"}
USE_SEG_MASTER = (len(SEG_RULES) > 0 and
                  any(c.lower() not in _EMPTY_CRITERIA for _, c in SEG_RULES))

if USE_SEG_MASTER:
    log(2, f"✅ FIX 2 — Business Segment Master wired | {len(SEG_RULES)} rules (dynamic)")
    audit(2, "segment_source", "Business Segment Master (dynamic)")
else:
    log(2, "⚠️  Business Segment Master — no criteria — hardcoded fallback")
    audit(2, "segment_source", "Hardcoded fallback")


# ── FIX 3 — Value Proposition Master ─────────────────────────────────────────
vp_master = raw_valueprops.copy()
vp_master.columns = ["tier_name", "risk_level", "action_type",
                      "message_template", "discount_pct", "channel", "priority"]
vp_master = vp_master[vp_master["tier_name"].notna() & vp_master["risk_level"].notna()].copy()
vp_master = vp_master[~vp_master["tier_name"].astype(str).str.startswith("──")]
vp_master["tier_name"]    = vp_master["tier_name"].astype(str).str.strip()
vp_master["risk_level"]   = vp_master["risk_level"].astype(str).str.strip()
vp_master["discount_pct"] = pd.to_numeric(vp_master["discount_pct"], errors="coerce").fillna(0)
vp_master["priority"]     = pd.to_numeric(vp_master["priority"],     errors="coerce").fillna(99)
vp_master = vp_master.sort_values("priority").reset_index(drop=True)

VP_LOOKUP = {}
for _, row in vp_master.iterrows():
    key = (row["tier_name"], row["risk_level"])
    if key not in VP_LOOKUP:
        VP_LOOKUP[key] = {
            "recommended_action": str(row["action_type"]),
            "message_template":   str(row["message_template"]),
            "discount_pct":       float(row["discount_pct"]),
            "preferred_channel":  str(row["channel"]),
        }

USE_VP_MASTER = len(VP_LOOKUP) > 0
if USE_VP_MASTER:
    log(2, f"✅ FIX 3 — Value Proposition Master wired | "
           f"{len(VP_LOOKUP)} (tier × risk) pairs (dynamic)")
    audit(2, "vp_source",  "Value Proposition Master (dynamic)")
    audit(2, "vp_entries", str(len(VP_LOOKUP)))
else:
    log(2, "⚠️  Value Proposition Master empty — default VP will be applied")
    audit(2, "vp_source", "Default fallback")

log(2, "Stage 2 complete")


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 3 — VALIDATE
# ══════════════════════════════════════════════════════════════════════════════
log(3, "Validating input data")

quarantine = []

cust_before = len(raw_customers)
raw_customers = raw_customers.dropna(subset=["customer_email", "customer_id"])
cust_dropped  = cust_before - len(raw_customers)
if cust_dropped:
    quarantine.append({"sheet": "Customer Master",
                       "reason": "missing email or customer_id", "count": cust_dropped})

dupes_email = raw_customers.duplicated("customer_email", keep="first").sum()
if dupes_email:
    quarantine.append({"sheet": "Customer Master",
                       "reason": "duplicate customer_email", "count": dupes_email})
raw_customers = raw_customers.drop_duplicates("customer_email", keep="first")

orders_before = len(raw_orders)
raw_orders = raw_orders.dropna(subset=["order_id", "customer_id", "order_date", "order_value_usd"])
ord_dropped = orders_before - len(raw_orders)
if ord_dropped:
    quarantine.append({"sheet": "Order Master",
                       "reason": "missing key fields", "count": ord_dropped})

neg_orders = (raw_orders["order_value_usd"] <= 0).sum()
if neg_orders:
    quarantine.append({"sheet": "Order Master",
                       "reason": "order_value_usd <= 0", "count": int(neg_orders)})
raw_orders = raw_orders[raw_orders["order_value_usd"] > 0]

# v5: product_id in Line Items is now INTEGER (was PROD-G007 string in v4)
li_before = len(raw_lineitems)
raw_lineitems = raw_lineitems.dropna(subset=["order_id", "product_id", "quantity"])
li_dropped = li_before - len(raw_lineitems)
if li_dropped:
    quarantine.append({"sheet": "Line Items",
                       "reason": "missing key fields", "count": li_dropped})

# Coerce product_id to int (v5 numeric PKs)
raw_lineitems["product_id"] = pd.to_numeric(raw_lineitems["product_id"],
                                             errors="coerce").fillna(0).astype(int)

audit(3, "customers_after_validation", len(raw_customers))
audit(3, "orders_after_validation",    len(raw_orders))
audit(3, "quarantined_rows",           sum(q["count"] for q in quarantine))
log(3, f"Validation complete — quarantined {sum(q['count'] for q in quarantine)} rows")


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 4 — NORMALISE
# ══════════════════════════════════════════════════════════════════════════════
log(4, "Normalising data types and formats")

raw_orders["order_date"] = pd.to_datetime(raw_orders["order_date"], errors="coerce")
raw_customers["account_created_date"] = pd.to_datetime(
    raw_customers["account_created_date"], errors="coerce")

for col in ["order_value_usd", "discount_usd", "order_item_count"]:
    raw_orders[col] = pd.to_numeric(raw_orders[col], errors="coerce").fillna(0)

for col in ["quantity", "unit_price_usd", "final_line_total_usd", "item_discount_usd"]:
    raw_lineitems[col] = pd.to_numeric(raw_lineitems[col], errors="coerce").fillna(0)

raw_customers["customer_email"] = raw_customers["customer_email"].str.lower().str.strip()
raw_orders["customer_id"]       = raw_orders["customer_id"].astype(str).str.strip()

for col in ["email_opt_in", "sms_opt_in"]:
    if col in raw_customers.columns:
        raw_customers[col] = (raw_customers[col].astype(str).str.lower()
                               .isin(["true", "1", "yes", "y"]).astype(int))

raw_orders["order_status"] = raw_orders["order_status"].str.strip().str.title()

future = (raw_orders["order_date"] > REFERENCE_DATE).sum()
if future:
    quarantine.append({"sheet": "Order Master",
                       "reason": "order_date in future", "count": int(future)})
raw_orders = raw_orders[raw_orders["order_date"] <= REFERENCE_DATE]

log(4, f"Normalisation complete — reference date: {REFERENCE_DATE.date()}")
audit(4, "future_orders_dropped", future)


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 5 — IDENTITY
#
# v5 composite key: (client_id, customer_id)
#   client_id  = the retail client (e.g. CLT-001 = Walmart) — from Vendor Config
#   customer_id = unique within that client's dataset
#   client_id already present in Customer Master / Order Master / Line Items in v5
# ══════════════════════════════════════════════════════════════════════════════
log(5, f"Building identity index — composite key: ({CLIENT_ID}, customer_id)")

# Stamp client_code and validate client_id on Customer Master
raw_customers["client_code"] = CLIENT_CODE
# In v5 the client_id column already exists; back-fill in case of nulls
if "client_id" in raw_customers.columns:
    raw_customers["client_id"] = raw_customers["client_id"].fillna(CLIENT_ID)
else:
    raw_customers["client_id"] = CLIENT_ID

email_to_cid = dict(zip(raw_customers["customer_email"], raw_customers["customer_id"]))
cid_to_email = {v: k for k, v in email_to_cid.items()}

norm_orders = raw_orders.merge(
    raw_customers[["customer_id", "client_id", "client_code",
                   "customer_email", "customer_name"]],
    on="customer_id", how="left"
)

raw_lineitems["customer_email"] = raw_lineitems["customer_id"].astype(str).map(cid_to_email)
raw_lineitems["client_code"]    = CLIENT_CODE

orphan_orders    = norm_orders["customer_email"].isna().sum()
orphan_lineitems = raw_lineitems["customer_email"].isna().sum()
if orphan_orders:
    quarantine.append({"sheet": "Order Master",
                       "reason": "no matching customer_id", "count": int(orphan_orders)})
if orphan_lineitems:
    quarantine.append({"sheet": "Line Items",
                       "reason": "no matching customer_id", "count": int(orphan_lineitems)})

norm_orders   = norm_orders.dropna(subset=["customer_email"])
raw_lineitems = raw_lineitems.dropna(subset=["customer_email"])

log(5, f"Identity resolved — {len(email_to_cid)} unique customers "
       f"| composite key: ({CLIENT_ID}, customer_id)")
audit(5, "client_id",          CLIENT_ID)
audit(5, "client_code",        CLIENT_CODE)
audit(5, "unique_customers",   len(email_to_cid))
audit(5, "orphaned_orders",    orphan_orders)


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 6 — REPEAT ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
log(6, "Computing repeat customer metrics and retention labels")

completed  = norm_orders[norm_orders["order_status"] == "Completed"].copy()
all_placed = norm_orders.copy()

txn_completed = (completed.groupby("customer_email")
                 .agg(
                     order_count_total   = ("order_id",        "nunique"),
                     first_order_date    = ("order_date",       "min"),
                     last_order_date     = ("order_date",       "max"),
                     total_gross_spend   = ("order_value_usd",  "sum"),
                     total_discount_usd  = ("discount_usd",     "sum"),
                     avg_order_value     = ("order_value_usd",  "mean"),
                     max_order_value     = ("order_value_usd",  "max"),
                     min_order_value     = ("order_value_usd",  "min"),
                     order_value_std     = ("order_value_usd",  "std"),
                     avg_items_per_order = ("order_item_count", "mean"),
                     coupon_used_count   = ("coupon_code",      lambda x: x.notna().sum()),
                 ).reset_index())

last_activity = (all_placed.groupby("customer_email")["order_date"]
                 .max().reset_index()
                 .rename(columns={"order_date": "last_activity_date"}))

acct_dates = raw_customers[["customer_email", "account_created_date"]].copy()

all_emails = pd.DataFrame({"customer_email": raw_customers["customer_email"].unique()})
txn = (all_emails
       .merge(txn_completed,  on="customer_email", how="left")
       .merge(last_activity,  on="customer_email", how="left")
       .merge(acct_dates,     on="customer_email", how="left"))

zero_fill_cols = ["order_count_total", "total_gross_spend", "total_discount_usd",
                  "avg_order_value",   "max_order_value",   "min_order_value",
                  "order_value_std",   "avg_items_per_order", "coupon_used_count"]
for col in zero_fill_cols:
    txn[col] = txn[col].fillna(0)

txn["days_since_last_order"] = (
    (REFERENCE_DATE - txn["last_activity_date"]).dt.days
    .fillna((REFERENCE_DATE - txn["account_created_date"]).dt.days)
    .fillna(999).astype(int)
)
txn["total_net_spend"]  = txn["total_gross_spend"] - txn["total_discount_usd"]
txn["account_age_days"] = ((REFERENCE_DATE - txn["account_created_date"])
                            .dt.days.fillna(0).astype(int))
txn["order_value_std"]  = txn["order_value_std"].fillna(0)

txn["all_orders_failed"] = (
    (txn["order_count_total"] == 0) &
    txn["customer_email"].isin(all_placed["customer_email"])
).astype(int)

log(6, f"All {len(txn)} customers retained — "
       f"{int((txn['order_count_total']==0).sum())} with zero completed orders via left join")


def compute_gaps(email):
    """
    Returns (mean_gap, median_gap, recent_gap) for a customer.
    - mean_gap   : mean of ALL inter-order gaps (kept for ML / backwards compat)
    - median_gap : robust central tendency — not skewed by one large historical gap
    - recent_gap : mean of last 3 gaps — reflects current buying rhythm
    Together these prevent distortion from a single historical outlier:
      e.g. gaps [5, 5, 200, 5] → mean=53.75 (bad), median=5, recent=5 (both correct)
    """
    dates = sorted(completed[completed["customer_email"] == email]["order_date"].unique())
    if len(dates) < 2:
        return np.nan, np.nan, np.nan
    gaps = [(dates[i + 1] - dates[i]).days for i in range(len(dates) - 1)]
    return float(np.mean(gaps)), float(np.median(gaps)), float(np.mean(gaps[-3:]))


gap_results = txn["customer_email"].apply(compute_gaps)
txn["avg_order_gap_days"]    = gap_results.apply(lambda x: x[0])
txn["median_order_gap_days"] = gap_results.apply(lambda x: x[1])
txn["recent_order_gap_days"] = gap_results.apply(lambda x: x[2])

for window, col in [(30, "orders_last_30d"), (60, "orders_last_60d"), (90, "orders_last_90d")]:
    cutoff = REFERENCE_DATE - pd.Timedelta(days=window)
    counts = (completed[completed["order_date"] >= cutoff]
              .groupby("customer_email")["order_id"].nunique())
    txn[col] = txn["customer_email"].map(counts).fillna(0).astype(int)

txn["personal_rhythm_breached"] = (
    (txn["median_order_gap_days"].notna()) &
    (
        (txn["days_since_last_order"] > txn["median_order_gap_days"] * RHYTHM_MULT) |
        (txn["days_since_last_order"] > txn["recent_order_gap_days"] * RHYTHM_MULT)
    )
).astype(int)

txn["coupon_usage_rate"] = (txn["coupon_used_count"] /
                            txn["order_count_total"].replace(0, np.nan)).fillna(0).round(4)

pay_div = (completed.groupby("customer_email")["payment_method"].nunique()
           .reset_index().rename(columns={"payment_method": "payment_diversity"}))
txn = txn.merge(pay_div, on="customer_email", how="left")

pref_pay = (completed.groupby(["customer_email", "payment_method"]).size()
            .reset_index(name="cnt")
            .sort_values("cnt", ascending=False)
            .drop_duplicates("customer_email")[["customer_email", "payment_method"]]
            .rename(columns={"payment_method": "preferred_payment_method"}))
txn = txn.merge(pref_pay, on="customer_email", how="left")


def retention_label(row):
    n          = row["order_count_total"]
    d          = row["days_since_last_order"]
    median_gap = row["median_order_gap_days"]
    recent_gap = row["recent_order_gap_days"]
    at_risk_baseline = (
        min(median_gap, recent_gap)
        if pd.notna(median_gap) and pd.notna(recent_gap)
        else (median_gap if pd.notna(median_gap) else recent_gap)
    )
    if n < REPEAT_THRESHOLD:          return "New"
    if pd.notna(median_gap) and d > median_gap * REACT_MULT: return "Reactivated"
    if pd.notna(at_risk_baseline) and d > at_risk_baseline * RHYTHM_MULT: return "At-Risk"
    if d >= CHURN_DAYS:               return "At-Risk"
    return "Returning"


txn["retention_label"] = txn.apply(retention_label, axis=1)

txn["is_churned"] = (
    txn["all_orders_failed"].astype(bool) |
    (
        (txn["personal_rhythm_breached"].astype(bool) |
         (txn["days_since_last_order"] >= CHURN_DAYS)) &
        (txn["order_count_total"] > 1) &
        ~txn["retention_label"].isin(["Returning", "Reactivated"])
    )
).astype(int)

log(6, f"Repeat analysis done — churned: {txn['is_churned'].sum()} / {len(txn)}")
audit(6, "churned_customers",     int(txn["is_churned"].sum()))
audit(6, "retention_New",         int((txn["retention_label"] == "New").sum()))
audit(6, "retention_Returning",   int((txn["retention_label"] == "Returning").sum()))
audit(6, "retention_AtRisk",      int((txn["retention_label"] == "At-Risk").sum()))
audit(6, "retention_Reactivated", int((txn["retention_label"] == "Reactivated").sum()))


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 7 — HIGH VALUE TIERS
# ══════════════════════════════════════════════════════════════════════════════
log(7, f"Assigning value tiers — method: {TIER_METHOD}  |  names: {TIER_NAMES}")

spend = txn["total_net_spend"]

if TIER_METHOD == "quartile":
    _cutpoints = []
    for q in TIER_THRESHOLDS:
        q_val = float(q)
        _cutpoints.append(spend.quantile(q_val) if q_val > 0 else -np.inf)

    def assign_tier(s):
        for tier_name, cutpoint in zip(TIER_NAMES, _cutpoints):
            if s >= cutpoint:
                return tier_name
        return TIER_NAMES[-1]

    for tname, q, cp in zip(TIER_NAMES, TIER_THRESHOLDS, _cutpoints):
        if cp != -np.inf:
            log(7, f"  {tname}: net_spend >= {cp:.2f}  (Q{float(q)*100:.0f})")
            audit(7, f"quartile_{tname}_min_spend", round(cp, 2))
else:
    def assign_tier(s):
        for tier_name, threshold in zip(TIER_NAMES, TIER_THRESHOLDS):
            if s >= float(threshold):
                return tier_name
        return TIER_NAMES[-1]

    for tname, thresh in zip(TIER_NAMES, TIER_THRESHOLDS):
        log(7, f"  {tname}: net_spend >= {thresh}")
        audit(7, f"fixed_{tname}_min_spend", thresh)

txn["value_tier"]    = txn["total_net_spend"].apply(assign_tier)
txn["is_high_value"] = (txn["value_tier"] == TIER_NAMES[0]).astype(int)
tier_weight_map = {name: len(TIER_NAMES) - i for i, name in enumerate(TIER_NAMES)}
txn["tier_weight"] = txn["value_tier"].map(tier_weight_map).fillna(1).astype(int)

log(7, "Tier distribution: " +
    " | ".join(f"{t}:{int((txn['value_tier']==t).sum())}" for t in TIER_NAMES))
audit(7, "tier_distribution", str(txn["value_tier"].value_counts().to_dict()))


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 8 — PRODUCT AGGREGATIONS + VENDOR FEATURES + BUSINESS SEGMENTS
#
# v3 additions:
#   - vendor_diversity   : # distinct product vendors behind the customer's purchases
#   - top_vendor         : vendor_name with most purchase spend
#   - brand_active_pct   : share of completed line items from active (onboarded) brands
#   - active_brand_purchases : raw count of active-brand line items
# ══════════════════════════════════════════════════════════════════════════════
log(8, "Computing product affinity, vendor features, RFM scores, and business segments")

li_enr = raw_lineitems.merge(
    product_enriched[["product_id", "category_name", "sub_category", "sub_sub_category",
                       "brand", "brand_active", "base_price_usd",
                       "product_vendor"]],
    on="product_id", how="left")
li_enr = li_enr.merge(
    raw_orders[["order_id", "order_status"]],
    on="order_id", how="left")
li_comp = li_enr[li_enr["item_status"] == "Completed"].copy()

# — Category aggregations ─────────────────────────────────────────────────────
cat_agg = (li_comp.groupby(["customer_email", "category_name"])
           .agg(cat_spend=("final_line_total_usd", "sum"),
                cat_qty=("quantity",              "sum"))
           .reset_index())

top_cat = (cat_agg.sort_values("cat_spend", ascending=False)
           .groupby("customer_email").first()
           .reset_index()[["customer_email", "category_name"]]
           .rename(columns={"category_name": "top_category"}))

cust_total_spend = cat_agg.groupby("customer_email")["cat_spend"].sum()
cat_agg["spend_share"] = cat_agg["cat_spend"] / cat_agg["customer_email"].map(cust_total_spend)
top_share = (cat_agg.sort_values("spend_share", ascending=False)
             .groupby("customer_email").first()
             .reset_index()[["customer_email", "spend_share"]]
             .rename(columns={"spend_share": "top_category_spend_share"}))

cat_div = (cat_agg.groupby("customer_email")["category_name"].nunique()
           .reset_index().rename(columns={"category_name": "category_diversity"}))
sub_div = (li_comp.groupby("customer_email")["sub_category"].nunique()
           .reset_index().rename(columns={"sub_category": "sub_category_diversity"}))
sub_sub_div = (li_comp.groupby("customer_email")["sub_sub_category"].nunique()
               .reset_index().rename(columns={"sub_sub_category": "sub_sub_category_diversity"}))

top_brand = (li_comp.groupby(["customer_email", "brand"])["final_line_total_usd"].sum()
             .reset_index().sort_values("final_line_total_usd", ascending=False)
             .drop_duplicates("customer_email")[["customer_email", "brand"]]
             .rename(columns={"brand": "top_brand"}))

brand_div = (li_comp.groupby("customer_email")["brand"].nunique()
             .reset_index().rename(columns={"brand": "brand_diversity"}))

avg_unit = (li_comp.groupby("customer_email")["unit_price_usd"].mean()
            .reset_index().rename(columns={"unit_price_usd": "avg_unit_price"}))

avg_qty = (li_comp.groupby("customer_email")["quantity"].mean()
           .reset_index().rename(columns={"quantity": "avg_qty_per_item"}))

# — Return / cancellation rates ───────────────────────────────────────────────
all_li = li_enr.copy()
ret_cnt = (all_li[all_li["item_status"] == "Returned"]
           .groupby("customer_email").size().reset_index(name="returned_items"))
all_cnt = (all_li.groupby("customer_email").size().reset_index(name="total_items"))
ret_rate = ret_cnt.merge(all_cnt, on="customer_email", how="right")
ret_rate["return_rate"]    = (ret_rate["returned_items"].fillna(0) /
                               ret_rate["total_items"].fillna(1)).round(4)
ret_rate["returned_items"] = ret_rate["returned_items"].fillna(0).astype(int)

cancel_cnt = (norm_orders[norm_orders["order_status"] == "Cancelled"]
              .groupby("customer_email").size().reset_index(name="cancelled_orders"))
total_ords  = (norm_orders.groupby("customer_email").size()
               .reset_index(name="total_orders_all"))
canc_rate = cancel_cnt.merge(total_ords, on="customer_email", how="right")
canc_rate["cancellation_rate"] = (canc_rate["cancelled_orders"].fillna(0) /
                                   canc_rate["total_orders_all"].fillna(1)).round(4)
canc_rate["cancelled_orders"]  = canc_rate["cancelled_orders"].fillna(0).astype(int)

# — v3 NEW: Vendor features ───────────────────────────────────────────────────
# vendor_diversity: how many distinct product vendors behind this customer's purchases
vend_div = (li_enr.groupby("customer_email")["product_vendor"]
            .nunique().reset_index()
            .rename(columns={"product_vendor": "vendor_diversity"}))

# top_vendor: vendor with the highest completed purchase spend per customer
top_vendor = (li_comp.groupby(["customer_email", "product_vendor"])["final_line_total_usd"]
              .sum().reset_index()
              .sort_values("final_line_total_usd", ascending=False)
              .drop_duplicates("customer_email")[["customer_email", "product_vendor"]]
              .rename(columns={"product_vendor": "top_vendor"}))

# brand_active_pct: share of completed line items from active brands
li_brand_active = li_comp.copy()
li_brand_active["brand_active"] = li_brand_active["brand_active"].fillna(0)
b_act_num = (li_brand_active[li_brand_active["brand_active"] == 1]
             .groupby("customer_email").size().reset_index(name="active_brand_purchases"))
b_act_tot = (li_brand_active.groupby("customer_email").size()
             .reset_index(name="total_brand_line_items"))
brand_active_pct_df = b_act_num.merge(b_act_tot, on="customer_email", how="right")
brand_active_pct_df["brand_active_pct"] = (
    brand_active_pct_df["active_brand_purchases"].fillna(0) /
    brand_active_pct_df["total_brand_line_items"].fillna(1)
).round(4)
brand_active_pct_df["active_brand_purchases"] = (
    brand_active_pct_df["active_brand_purchases"].fillna(0).astype(int))
brand_active_pct_df = brand_active_pct_df[
    ["customer_email", "brand_active_pct", "active_brand_purchases"]]

log(8, f"v3 vendor features — median vendor_diversity: "
       f"{vend_div['vendor_diversity'].median():.1f}  |  "
       f"top_vendors found: {top_vendor['top_vendor'].notna().sum()}")

# — RFM Scores (quintile bins 1–5) ────────────────────────────────────────────
def rfm_score(series, ascending=False):
    labels = [1, 2, 3, 4, 5] if ascending else [5, 4, 3, 2, 1]
    try:
        return pd.qcut(series, 5, labels=labels, duplicates="drop").astype(int)
    except Exception:
        return pd.cut(series, 5, labels=labels, duplicates="drop").astype(int)


txn["recency_score"]   = rfm_score(txn["days_since_last_order"], ascending=False)
txn["frequency_score"] = rfm_score(txn["order_count_total"],     ascending=True)
txn["monetary_score"]  = rfm_score(txn["total_net_spend"],       ascending=True)

zero_order_mask = txn["order_count_total"] == 0
txn.loc[zero_order_mask, "recency_score"]   = 1
txn.loc[zero_order_mask, "frequency_score"] = 1
txn.loc[zero_order_mask, "monetary_score"]  = 1
n_clamped = int(zero_order_mask.sum())
log(8, f"✅ Zero-order clamp — {n_clamped} customers clamped to RFM 1+1+1=3")
audit(8, "champions_fix_customers_clamped", n_clamped)

txn["rfm_total_score"] = (txn["recency_score"] +
                           txn["frequency_score"] +
                           txn["monetary_score"])

# — Business Segment Assignment ───────────────────────────────────────────────
def assign_segment_dynamic(row):
    namespace = {
        "recency_score":         int(row.get("recency_score",         0)),
        "frequency_score":       int(row.get("frequency_score",       0)),
        "monetary_score":        int(row.get("monetary_score",        0)),
        "rfm_total_score":       int(row.get("rfm_total_score",       0)),
        "days_since_last_order": int(row.get("days_since_last_order", 999)),
        "order_count_total":     int(row.get("order_count_total",     0)),
        "total_net_spend":       float(row.get("total_net_spend",     0)),
        "is_churned":            int(row.get("is_churned",            0)),
    }
    for seg_name, criteria in SEG_RULES:
        crit = str(criteria).strip()
        if crit.lower() in _EMPTY_CRITERIA:
            return seg_name
        try:
            if eval(crit, {"__builtins__": {}}, namespace):
                return seg_name
        except Exception:
            continue
    return "Unknown"


def assign_segment_hardcoded(row):
    r, f, m, total = (row["recency_score"], row["frequency_score"],
                      row["monetary_score"], row["rfm_total_score"])
    if total >= 12:             return "Champions"
    if f >= 4 and m >= 3:       return "Loyal Customers"
    if r >= 4 and 2 <= f <= 3:  return "Potential Loyalists"
    if r <= 2 and f >= 3:       return "At-Risk Customers"
    if r >= 4 and f <= 2:       return "New Customers"
    if row["days_since_last_order"] > 180 and r == 1: return "Lost"
    if total <= 5:              return "Hibernating"
    return "Potential Loyalists"


if USE_SEG_MASTER:
    txn["business_segment"] = txn.apply(assign_segment_dynamic, axis=1)
    log(8, "✅ Business segments from Business Segment Master (dynamic)")
    audit(8, "segment_assignment_method", "Business Segment Master (dynamic)")
else:
    txn["business_segment"] = txn.apply(assign_segment_hardcoded, axis=1)
    log(8, "⚠️  Business segments from hardcoded rules (master fallback)")
    audit(8, "segment_assignment_method", "Hardcoded fallback")

log(8, "Segment distribution: " +
    " | ".join(f"{s}:{n}" for s, n in txn["business_segment"].value_counts().items()))
audit(8, "categories_covered",     int(li_comp["category_name"].nunique()))
audit(8, "unique_brands",          int(li_comp["brand"].nunique()))
audit(8, "avg_category_diversity", round(cat_div["category_diversity"].mean(), 2))
audit(8, "unique_vendors_in_purchases", int(li_enr["product_vendor"].nunique()))


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 9 — FEATURE ENGINEERING + VALUE PROPOSITION LOOKUP
# ══════════════════════════════════════════════════════════════════════════════
log(9, "Building ML feature vector")

feat = txn.copy()

for df_merge in [
    top_cat,    top_share,  cat_div,    sub_div,    sub_sub_div,
    top_brand,  brand_div,  avg_unit,   avg_qty,
    ret_rate[["customer_email", "return_rate", "returned_items"]],
    canc_rate[["customer_email", "cancellation_rate", "cancelled_orders"]],
    # v3: vendor features
    vend_div,   top_vendor, brand_active_pct_df,
]:
    feat = feat.merge(df_merge, on="customer_email", how="left")

# Merge in customer profile columns
feat = feat.merge(
    raw_customers[["customer_email", "customer_name", "customer_id",
                   "client_id", "client_code",
                   "account_created_date", "registration_channel",
                   "preferred_device", "email_opt_in", "sms_opt_in",
                   "state", "country_code"]],
    on="customer_email", how="left"
)
feat["client_id"]   = feat["client_id"].fillna(CLIENT_ID)
feat["client_code"] = feat["client_code"].fillna(CLIENT_CODE)

feat["discount_rate"]       = (feat["total_discount_usd"] /
                                feat["total_gross_spend"].replace(0, np.nan)).fillna(0).round(4)
feat["days_inactive_ratio"] = (feat["days_since_last_order"] /
                                feat["avg_order_gap_days"].replace(0, np.nan)).fillna(0).round(4)
feat["urgency_score"]       = ((feat["is_churned"] * feat["tier_weight"]) +
                                (feat["days_since_last_order"] / 180)).round(4)
feat["priority_score"]      = (feat["is_churned"] * feat["tier_weight"]).round(4)

num_cols = ["avg_order_gap_days", "median_order_gap_days", "recent_order_gap_days",
            "order_value_std", "category_diversity", "sub_category_diversity",
            "brand_diversity", "avg_unit_price", "avg_qty_per_item",
            "return_rate", "cancellation_rate", "payment_diversity",
            "top_category_spend_share", "vendor_diversity", "brand_active_pct",
            "active_brand_purchases"]
for col in num_cols:
    if col in feat.columns:
        feat[col] = feat[col].fillna(0)

feat["registration_channel_enc"] = feat["registration_channel"].astype("category").cat.codes
feat["preferred_device_enc"]     = feat["preferred_device"].astype("category").cat.codes
feat["top_category_enc"]         = feat["top_category"].astype("category").cat.codes
feat["top_brand_enc"]            = feat["top_brand"].astype("category").cat.codes
feat["top_vendor_enc"]           = feat["top_vendor"].astype("category").cat.codes  # v3 NEW

# — Value Proposition Lookup ───────────────────────────────────────────────────
def apply_vp_lookup(row):
    tier = str(row.get("value_tier",     ""))
    risk = str(row.get("retention_label", ""))
    for key in [(tier, risk), (tier, "All"), ("All", risk)]:
        if key in VP_LOOKUP:
            return pd.Series(VP_LOOKUP[key])
    return pd.Series({
        "recommended_action": "Standard Engagement",
        "message_template":   f"Thank you for shopping with {CLIENT_NAME}!",
        "discount_pct":       0.0,
        "preferred_channel":  "email",
    })


if USE_VP_MASTER:
    vp_applied = feat.apply(apply_vp_lookup, axis=1)
    feat       = pd.concat([feat, vp_applied], axis=1)
    matched_vp = sum(1 for _, r in feat.iterrows()
                     if (r["value_tier"], r["retention_label"]) in VP_LOOKUP)
    log(9, f"✅ VP lookup applied — {matched_vp}/{len(feat)} customers matched a rule "
           f"({len(feat) - matched_vp} used default)")
    audit(9, "vp_assignment_method",  "Value Proposition Master (dynamic)")
    audit(9, "vp_customers_matched",  matched_vp)
    audit(9, "vp_customers_default",  len(feat) - matched_vp)
else:
    feat["recommended_action"] = "Standard Engagement"
    feat["message_template"]   = f"Thank you for shopping with {CLIENT_NAME}!"
    feat["discount_pct"]       = 0.0
    feat["preferred_channel"]  = "email"
    log(9, "⚠️  VP Master empty — default propositions assigned")
    audit(9, "vp_assignment_method", "Default fallback")


# — Final feature vector ───────────────────────────────────────────────────────
IDENTITY_COLS = [
    # v3: client_id replaces vendor_id as the retail-client identifier
    "client_id",    # CLT-001 — scopes all data to this retail client
    "client_code",  # WMT — human-readable short code
    "customer_id",
    "customer_email", "customer_name",
    "value_tier", "business_segment", "retention_label", "is_churned",
    # VP recommendation
    "recommended_action", "message_template", "discount_pct", "preferred_channel",
]

FEATURE_COLS = [
    # RFM
    "days_since_last_order", "order_count_total", "total_net_spend",
    "recency_score", "frequency_score", "monetary_score", "rfm_total_score",
    # Order behaviour
    "avg_order_value", "avg_items_per_order", "max_order_value",
    "min_order_value", "order_value_std", "account_age_days",
    "avg_order_gap_days", "median_order_gap_days", "recent_order_gap_days",
    "orders_last_30d", "orders_last_60d", "orders_last_90d",
    # Financial
    "total_gross_spend", "total_discount_usd", "discount_rate",
    "coupon_usage_rate", "payment_diversity",
    # Product
    "category_diversity", "sub_category_diversity", "sub_sub_category_diversity",
    "top_category_spend_share", "avg_unit_price", "avg_qty_per_item", "brand_diversity",
    # v3 NEW: Vendor & brand availability features
    "vendor_diversity",       # distinct product vendors behind customer's purchases
    "brand_active_pct",       # share of purchases from active (onboarded) brands
    "active_brand_purchases", # raw count of active-brand line items
    # Returns / cancellations
    "return_rate", "returned_items", "cancellation_rate", "cancelled_orders",
    # Customer profile
    "email_opt_in", "sms_opt_in",
    # Churn signals
    "personal_rhythm_breached", "all_orders_failed",
    "days_inactive_ratio", "is_high_value", "tier_weight",
    "urgency_score", "priority_score",
    # Encoded categoricals
    "registration_channel_enc", "preferred_device_enc",
    "top_category_enc", "top_brand_enc", "top_vendor_enc",  # v3 NEW
]

available  = [c for c in IDENTITY_COLS + FEATURE_COLS if c in feat.columns]
ai_vector  = feat[available].reset_index(drop=True)

log(9, f"Feature vector ready — {len(ai_vector)} customers × {len(available)} columns")
audit(9, "customers_in_ml_vector", len(ai_vector))
audit(9, "total_features",         len(available))
audit(9, "churn_rate_pct",         round(ai_vector["is_churned"].mean() * 100, 2))
audit(9, "new_v3_features",        "vendor_diversity, brand_active_pct, active_brand_purchases, top_vendor_enc")


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 10 — OUTPUT  (write clean_ai_dataset_v7.xlsx — 12 sheets)
# ══════════════════════════════════════════════════════════════════════════════
log(10, f"Writing output: {OUTPUT_FILE}")

wb = Workbook()
wb.remove(wb.active)

# ── Sheet 1: AI Feature Vector ────────────────────────────────────────────────
write_sheet(wb, "🧠 AI Feature Vector", ai_vector,
    header_color="1A237E", tab_color="1A237E",
    freeze="D3",
    banner_subtitle=(f"{len(ai_vector)} rows  |  {len(available)} features  |  "
                     f"Target: is_churned  |  v3: vendor_diversity + brand_active_pct added  |  "
                     f"Client: {CLIENT_NAME}  ({CLIENT_ID})"))

# ── Sheet 2: Repeat Analysis ──────────────────────────────────────────────────
repeat_cols = ["client_id", "client_code", "customer_email", "customer_name", "customer_id",
               "order_count_total", "first_order_date", "last_order_date",
               "days_since_last_order", "account_age_days",
               "avg_order_gap_days", "median_order_gap_days", "recent_order_gap_days",
               "orders_last_30d", "orders_last_60d", "orders_last_90d",
               "personal_rhythm_breached", "retention_label", "is_churned",
               "coupon_usage_rate", "preferred_payment_method"]
repeat_df = feat[[c for c in repeat_cols if c in feat.columns]].copy()
write_sheet(wb, "🔁 Repeat Analysis", repeat_df,
    header_color="2E7D32", tab_color="2E7D32",
    freeze="D3",
    banner_subtitle=(f"{len(repeat_df)} rows  |  Order frequency, gaps, retention  |  "
                     f"Churn threshold: {CHURN_DAYS} days  |  Client: {CLIENT_NAME}"))

# ── Sheet 3: High Value Classification ────────────────────────────────────────
hv_cols = ["client_id", "client_code", "customer_email", "customer_name", "customer_id",
           "total_gross_spend", "total_discount_usd", "total_net_spend",
           "value_tier", "tier_weight", "is_high_value",
           "avg_order_value", "max_order_value", "order_count_total",
           "recency_score", "monetary_score", "business_segment"]
hv_df = feat[[c for c in hv_cols if c in feat.columns]].sort_values(
    "total_net_spend", ascending=False)
write_sheet(wb, "💎 High Value", hv_df,
    header_color="B71C1C", tab_color="B71C1C",
    freeze="D3",
    banner_subtitle=(f"Tier method: {TIER_METHOD.upper()}  |  "
                     f"Tiers: {' > '.join(TIER_NAMES)}  |  Client: {CLIENT_NAME}"))

# ── Sheet 4: Normalised Orders ────────────────────────────────────────────────
if "client_id" not in norm_orders.columns:
    norm_orders["client_id"] = CLIENT_ID
norm_out = norm_orders[["client_id", "order_id", "customer_id", "customer_email",
                          "customer_name", "order_date", "order_status",
                          "order_value_usd", "discount_usd", "coupon_code",
                          "payment_method", "order_item_count"]].copy()
write_sheet(wb, "📦 Normalised Orders", norm_out,
    header_color="1565C0", tab_color="1565C0",
    freeze="E3",
    banner_subtitle=(f"{len(norm_out)} rows  |  All valid orders after validation  |  "
                     f"Client: {CLIENT_NAME}"))

# ── Sheet 5: RFM Analysis ─────────────────────────────────────────────────────
rfm_cols = ["client_id", "client_code", "customer_email", "customer_name", "customer_id",
            "days_since_last_order", "order_count_total", "total_net_spend",
            "recency_score", "frequency_score", "monetary_score", "rfm_total_score",
            "business_segment"]
rfm_df = feat[[c for c in rfm_cols if c in feat.columns]].sort_values(
    "rfm_total_score", ascending=False)
write_sheet(wb, "📈 RFM Analysis", rfm_df,
    header_color="6A1B9A", tab_color="6A1B9A",
    freeze="D3",
    banner_subtitle=(f"RFM scores 1–5  |  Max: 15  |  "
                     f"Segments: {'Business Segment Master' if USE_SEG_MASTER else 'hardcoded'}  |  "
                     f"Client: {CLIENT_NAME}"))

# ── Sheet 6: Product Affinity ─────────────────────────────────────────────────
pa_cols = ["client_id", "client_code", "customer_email", "customer_name", "customer_id",
           "top_category", "top_category_spend_share", "category_diversity",
           "sub_category_diversity", "sub_sub_category_diversity",
           "top_brand", "brand_diversity", "brand_active_pct",
           "top_vendor", "vendor_diversity",    # v3: vendor columns
           "avg_unit_price", "avg_qty_per_item",
           "return_rate", "cancellation_rate", "value_tier"]
pa_df = feat[[c for c in pa_cols if c in feat.columns]]
write_sheet(wb, "🛍️ Product Affinity", pa_df,
    header_color="E65100", tab_color="E65100",
    freeze="D3",
    banner_subtitle=(f"Category spend shares + vendor affinity  |  "
                     f"v3: brand_active_pct + vendor_diversity added  |  "
                     f"Client: {CLIENT_NAME}"))

# ── Sheet 7: Quarantine Log ───────────────────────────────────────────────────
qlog_df = (pd.DataFrame(quarantine) if quarantine
           else pd.DataFrame(columns=["sheet", "reason", "count"]))
ws_q = wb.create_sheet("⚠️ Quarantine Log")
ws_q.sheet_properties.tabColor = "F57F17"
ws_q.merge_cells("A1:D1")
bc = ws_q.cell(1, 1, (f"⚠️ Quarantine Log   |   "
                       f"{qlog_df['count'].sum() if len(qlog_df) else 0} rows quarantined  |  "
                       f"Client: {CLIENT_NAME}"))
bc.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
bc.fill = PatternFill("solid", start_color="F57F17")
bc.alignment = Alignment(horizontal="left", vertical="center")
for c, h in enumerate(["Sheet", "Reason", "Count", "Status"], 1):
    cell = ws_q.cell(2, c, h)
    cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="E65100")
    cell.border = thin_border()
for r, row in enumerate(qlog_df.itertuples(index=False), 3):
    for c, val in enumerate(list(row) + ["Quarantined"], 1):
        cell = ws_q.cell(r, c, val)
        cell.font = Font(name="Arial", size=9)
        cell.border = thin_border()
if len(qlog_df) == 0:
    ws_q.cell(3, 1, "✅ No rows quarantined — all records passed validation")
ws_q.column_dimensions["A"].width = 20
ws_q.column_dimensions["B"].width = 45
ws_q.column_dimensions["C"].width = 12
ws_q.column_dimensions["D"].width = 16

# ── Sheet 8: Pipeline Audit Log ──────────────────────────────────────────────
audit_df = pd.DataFrame(audit_log)
write_sheet(wb, "🔍 Pipeline Audit", audit_df,
    header_color="37474F", tab_color="37474F",
    banner_subtitle=(f"Stage-by-stage run summary  |  "
                     f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                     f"Client: {CLIENT_NAME}  (v3 pipeline)"))

# ── Sheet 9: Customer Category Spend ─────────────────────────────────────────
cat_spend_out = (li_comp.groupby(["customer_id", "customer_email",
                                   "category_name", "sub_category", "sub_sub_category"])
                 .agg(total_spend_usd=("final_line_total_usd", "sum"),
                      total_quantity=("quantity",              "sum"))
                 .reset_index())
cust_tot_map = cat_spend_out.groupby("customer_email")["total_spend_usd"].transform("sum")
cat_spend_out["pct_of_total_spend"] = (cat_spend_out["total_spend_usd"] /
                                        cust_tot_map * 100).round(2)
cat_spend_out["total_spend_usd"] = cat_spend_out["total_spend_usd"].round(2)
cat_spend_out = cat_spend_out.sort_values(["customer_email", "total_spend_usd"],
                                           ascending=[True, False])
cat_spend_out["spend_rank"] = cat_spend_out.groupby("customer_email").cumcount() + 1
cat_spend_out = cat_spend_out.merge(
    feat[["customer_email", "customer_name"]].drop_duplicates(),
    on="customer_email", how="left")
cat_spend_out["client_id"] = CLIENT_ID
cat_spend_out = cat_spend_out[["client_id", "customer_id", "customer_email", "customer_name",
                                "spend_rank", "category_name", "sub_category", "sub_sub_category",
                                "total_spend_usd", "total_quantity", "pct_of_total_spend"]]
write_sheet(wb, "🛒 Customer Category Spend", cat_spend_out,
    header_color="1565C0", tab_color="0277BD",
    freeze="E3",
    banner_subtitle=(f"{len(cat_spend_out)} rows  |  Per-customer category spend (desc)  |  "
                     f"Client: {CLIENT_NAME}"),
    col_widths={"A": 14, "B": 18, "C": 30, "D": 24, "E": 12, "F": 22,
                "G": 22, "H": 26, "I": 20, "J": 16, "K": 22})

# ── Sheet 10: Client Sales Dashboard ─────────────────────────────────────────
sales_out = (li_comp.groupby(["category_name", "sub_category", "sub_sub_category"])
             .agg(total_quantity_sold=("quantity",           "sum"),
                  total_revenue_usd=("final_line_total_usd", "sum"),
                  num_customers=("customer_id",              "nunique"))
             .reset_index())
ret_li_out  = li_enr[li_enr["item_status"] == "Returned"]
ret_cnt_out = (ret_li_out.groupby(["category_name", "sub_category", "sub_sub_category"])["quantity"]
               .sum().reset_index().rename(columns={"quantity": "returned_qty"}))
all_qty_out = (li_enr.groupby(["category_name", "sub_category", "sub_sub_category"])["quantity"]
               .sum().reset_index().rename(columns={"quantity": "total_all_qty"}))
sales_out = sales_out.merge(ret_cnt_out, on=["category_name", "sub_category", "sub_sub_category"], how="left")
sales_out = sales_out.merge(all_qty_out, on=["category_name", "sub_category", "sub_sub_category"], how="left")
sales_out["returned_qty"]      = sales_out["returned_qty"].fillna(0)
sales_out["return_rate_pct"]   = (sales_out["returned_qty"] /
                                   sales_out["total_all_qty"] * 100).round(2)
sales_out["avg_unit_price"]    = (sales_out["total_revenue_usd"] /
                                   sales_out["total_quantity_sold"]).round(2)
sales_out["total_revenue_usd"] = sales_out["total_revenue_usd"].round(2)
sd_r = sales_out.sort_values("total_revenue_usd", ascending=False).copy()
sd_r["revenue_rank"] = range(1, len(sd_r) + 1)
sd_q = sales_out.sort_values("total_quantity_sold", ascending=False).copy()
sd_q["quantity_rank"] = range(1, len(sd_q) + 1)
sales_out = sd_r.merge(sd_q[["category_name", "sub_category", "sub_sub_category", "quantity_rank"]],
                        on=["category_name", "sub_category", "sub_sub_category"])
sales_out = sales_out[["revenue_rank", "quantity_rank", "category_name", "sub_category",
                         "sub_sub_category", "total_quantity_sold", "total_revenue_usd",
                         "avg_unit_price", "num_customers", "return_rate_pct"]
                       ].sort_values("revenue_rank")
write_sheet(wb, "📊 Client Sales Dashboard", sales_out,
    header_color="BF360C", tab_color="BF360C",
    freeze="C3",
    banner_subtitle=(f"{len(sales_out)} category rows  |  Revenue + Quantity ranks  |  "
                     f"Client: {CLIENT_NAME}"),
    col_widths={"A": 16, "B": 16, "C": 22, "D": 22, "E": 28,
                "F": 22, "G": 22, "H": 18, "I": 16, "J": 18})

# ── Sheet 11: Vendor Analysis  (v3 NEW) ──────────────────────────────────────
# Per-vendor contribution: which suppliers drive the most revenue / customer reach
# li_comp and li_enr already have product_vendor from the Stage 8 merge; just
# add vendor_id (the numeric FK) so we can join brand counts later.
_vid_map = product_enriched[["product_id", "vendor_id"]].drop_duplicates("product_id")
li_vendor = li_comp.merge(_vid_map, on="product_id", how="left")
li_vendor_all = li_enr.merge(_vid_map, on="product_id", how="left")

vendor_sales = (li_vendor.groupby("product_vendor")
                .agg(total_revenue_usd=("final_line_total_usd", "sum"),
                     total_qty_sold=("quantity",               "sum"),
                     num_customers=("customer_id",             "nunique"),
                     num_products=("product_id",               "nunique"))
                .reset_index())

ret_vendor = (li_vendor_all[li_vendor_all["item_status"] == "Returned"]
              .groupby("product_vendor")["quantity"].sum()
              .reset_index().rename(columns={"quantity": "returned_qty"}))
all_vendor = (li_vendor_all.groupby("product_vendor")["quantity"].sum()
              .reset_index().rename(columns={"quantity": "total_all_qty"}))

vendor_sales = vendor_sales.merge(ret_vendor, on="product_vendor", how="left")
vendor_sales = vendor_sales.merge(all_vendor, on="product_vendor", how="left")
vendor_sales["returned_qty"]      = vendor_sales["returned_qty"].fillna(0)
vendor_sales["return_rate_pct"]   = (vendor_sales["returned_qty"] /
                                      vendor_sales["total_all_qty"] * 100).round(2)
vendor_sales["avg_unit_price"]    = (vendor_sales["total_revenue_usd"] /
                                      vendor_sales["total_qty_sold"].replace(0, np.nan)).round(2)
vendor_sales["total_revenue_usd"] = vendor_sales["total_revenue_usd"].round(2)
vendor_sales["revenue_share_pct"] = (vendor_sales["total_revenue_usd"] /
                                      vendor_sales["total_revenue_usd"].sum() * 100).round(2)

# Add brand count per vendor from Brand Master
brand_per_vendor = (raw_brands[raw_brands["active"] == 1]
                    .merge(raw_vendors[["vendor_id", "vendor_name"]].rename(
                           columns={"vendor_name": "product_vendor"}),
                           on="vendor_id", how="left")
                    .groupby("product_vendor")["brand_name"].nunique()
                    .reset_index().rename(columns={"brand_name": "active_brands"}))
vendor_sales = vendor_sales.merge(brand_per_vendor, on="product_vendor", how="left")
vendor_sales["active_brands"] = vendor_sales["active_brands"].fillna(0).astype(int)

vendor_sales = vendor_sales.sort_values("total_revenue_usd", ascending=False).reset_index(drop=True)
vendor_sales.insert(0, "vendor_rank", range(1, len(vendor_sales) + 1))
vendor_sales = vendor_sales.rename(columns={"product_vendor": "vendor_name"})
vendor_out = vendor_sales[["vendor_rank", "vendor_name", "active_brands",
                             "num_products", "num_customers",
                             "total_qty_sold", "total_revenue_usd",
                             "revenue_share_pct", "avg_unit_price", "return_rate_pct"]]

write_sheet(wb, "🏭 Vendor Analysis", vendor_out,
    header_color="4A148C", tab_color="4A148C",
    freeze="B3",
    banner_subtitle=(f"{len(vendor_out)} vendors  |  Revenue contribution per product supplier  |  "
                     f"v3 NEW  |  Client: {CLIENT_NAME}"),
    col_widths={"A": 14, "B": 24, "C": 16, "D": 16, "E": 16,
                "F": 18, "G": 20, "H": 20, "I": 18, "J": 18})

# ── Sheet 12: Action Plan ─────────────────────────────────────────────────────
action_cols = ["client_id", "client_code", "customer_email", "customer_name", "customer_id",
               "value_tier", "retention_label", "is_churned",
               "urgency_score", "priority_score",
               "recommended_action", "message_template",
               "discount_pct", "preferred_channel",
               "days_since_last_order", "total_net_spend",
               "business_segment", "top_vendor", "brand_active_pct"]  # v3: added vendor + brand fields
action_df = feat[[c for c in action_cols if c in feat.columns]].copy()
action_df = action_df.sort_values(["is_churned", "priority_score"],
                                   ascending=[False, False]).reset_index(drop=True)

write_sheet(wb, "💡 Action Plan", action_df,
    header_color="00695C", tab_color="00695C",
    freeze="D3",
    banner_subtitle=(f"{int(action_df['is_churned'].sum())} churned customers prioritised first  |  "
                     f"VP: {'Value Proposition Master (dynamic)' if USE_VP_MASTER else 'default fallback'}  |  "
                     f"v3: top_vendor + brand_active_pct added  |  Client: {CLIENT_NAME}"),
    col_widths={"A": 14, "B": 12, "C": 30, "D": 22, "E": 18, "F": 12,
                "G": 14, "H": 12, "I": 14, "J": 14, "K": 28,
                "L": 55, "M": 14, "N": 16, "O": 22, "P": 18, "Q": 26,
                "R": 22, "S": 18})

wb.save(OUTPUT_FILE)

# ── Final summary ──────────────────────────────────────────────────────────────
log(10, f"✅ Output saved → {OUTPUT_FILE}")
log(10, f"   Sheets ({len(wb.worksheets)}): {[s.title for s in wb.worksheets]}")
log(10, f"   Customers in ML vector  : {len(ai_vector)}")
log(10, f"   Total features          : {len(available)}")
log(10, f"   Churned customers       : {int(ai_vector['is_churned'].sum())} "
        f"({ai_vector['is_churned'].mean()*100:.1f}%)")
log(10, f"   Tier distribution       : " +
        " | ".join(f"{t}:{int((ai_vector['value_tier']==t).sum())}" for t in TIER_NAMES))
log(10, f"   Vendor coverage         : {len(vendor_out)} vendors analysed")
log(10, f"   Tier source             : "
        f"{'Value-Tier Master (dynamic)' if len(tier_master) >= 2 else 'Vendor Config (fallback)'}")
log(10, f"   Segment source          : "
        f"{'Business Segment Master (dynamic)' if USE_SEG_MASTER else 'Hardcoded fallback'}")
log(10, f"   VP source               : "
        f"{'Value Proposition Master (dynamic)' if USE_VP_MASTER else 'Default fallback'}")

print("\n" + "=" * 72)
print(f"  preprocess_walmart_v3.py  —  COMPLETE")
print(f"  Client  : {CLIENT_CODE}  ({CLIENT_NAME})")
print(f"  Input   : {INPUT_FILE}")
print(f"  Output  : {OUTPUT_FILE}")
print(f"  Schema  : v5  (client_id + Vendor Master + row-wise pricing)")
print(f"  Fixes   : ✅ client_id  ✅ Vendor Master  ✅ Product-Vendor Mapping")
print(f"            ✅ Row-wise price tiers  ✅ brand_active_pct  ✅ vendor_diversity")
print(f"            ✅ Tier Master  ✅ Segment Master  ✅ VP Master")
print("=" * 72)
